"""
Project Objects Router — FR-023, FR-024, FR-025, FR-026
ADR-004: JSONB flexible strategy + Pydantic discriminated union
Tables: ppg_project_objects, ppg_object_connections
"""
from __future__ import annotations

import io
import json
import re
from typing import Literal, Optional
from uuid import UUID, uuid4

import asyncpg
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, field_validator, model_validator

from app.auth import CurrentUser
from app.database import get_db
from app.services.audit_service import log_audit

router = APIRouter(prefix="/projects", tags=["project-objects"])

# ---------------------------------------------------------------------------
# Pydantic discriminated union — ADR-004
# ---------------------------------------------------------------------------

class WebAppInfo(BaseModel):
    object_type: Literal["web_app"]
    tech_stack: str = Field(..., description="e.g. React + FastAPI")
    version: str
    url_dev: Optional[str] = None
    url_staging: Optional[str] = None
    url_uat: Optional[str] = None
    url_prod: Optional[str] = None
    deployment_type: Optional[Literal["on-premise", "cloud", "hybrid"]] = None
    sso_enabled: bool = False
    notes: Optional[str] = None


class MobileAppInfo(BaseModel):
    object_type: Literal["mobile_app"]
    platform: Literal["iOS", "Android", "cross-platform"]
    version: str
    store_link_ios: Optional[str] = None
    store_link_android: Optional[str] = None
    tech_stack: Optional[str] = None
    min_os_version: Optional[str] = None
    notes: Optional[str] = None


class ApiInfo(BaseModel):
    object_type: Literal["api"]
    base_url: str
    auth_method: Literal["JWT", "OAuth2", "API_Key", "Basic", "None"]
    version: str
    url_dev: Optional[str] = None
    url_uat: Optional[str] = None
    protocol: Literal["REST", "SOAP", "GraphQL", "gRPC"] = "REST"
    endpoints: list[dict] = Field(default_factory=list)
    swagger_url: Optional[str] = None
    notes: Optional[str] = None


class EltInfo(BaseModel):
    object_type: Literal["elt"]
    source_system: str
    target_system: str
    schedule: str
    technology: Optional[str] = None
    data_format: Optional[str] = None
    volume_estimate: Optional[str] = None
    sla_minutes: Optional[int] = None
    notes: Optional[str] = None


OBJECT_TYPES = {"web_app", "mobile_app", "api", "elt"}
CODE_PATTERN = re.compile(r"^[A-Z0-9_]+$")

# ---------------------------------------------------------------------------
# Request / Response schemas
# ---------------------------------------------------------------------------

class ProjectObjectCreate(BaseModel):
    object_type: str
    name: str = Field(..., max_length=200)
    code: str = Field(..., max_length=50)
    description: Optional[str] = None
    owner: Optional[str] = Field(None, max_length=100)
    standard_info: dict = Field(..., description="Type-specific fields per ADR-004")

    @field_validator("object_type")
    @classmethod
    def validate_object_type(cls, v: str) -> str:
        if v not in OBJECT_TYPES:
            raise ValueError(f"object_type must be one of {sorted(OBJECT_TYPES)}")
        return v

    @field_validator("code")
    @classmethod
    def validate_code(cls, v: str) -> str:
        if not CODE_PATTERN.match(v):
            raise ValueError("code must be uppercase A-Z, 0-9, underscore only")
        return v

    @model_validator(mode="after")
    def validate_standard_info(self) -> "ProjectObjectCreate":
        """Validate standard_info per object_type using discriminated union."""
        info_with_type = {"object_type": self.object_type, **self.standard_info}
        # Use pydantic to validate the typed schema
        type_map: dict = {
            "web_app": WebAppInfo,
            "mobile_app": MobileAppInfo,
            "api": ApiInfo,
            "elt": EltInfo,
        }
        model_cls = type_map.get(self.object_type)
        if model_cls:
            model_cls.model_validate(info_with_type)
        return self


class ProjectObjectUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=200)
    description: Optional[str] = None
    owner: Optional[str] = Field(None, max_length=100)
    standard_info: Optional[dict] = None


class ConnectionCreate(BaseModel):
    target_object_id: UUID
    connection_type: Literal["api_call", "data_feed", "event", "file_transfer", "db_sync", "other"]
    protocol: Optional[str] = Field(None, max_length=30)
    frequency: Optional[str] = Field(None, max_length=50)
    description: Optional[str] = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_obj(row: asyncpg.Record) -> dict:
    d = dict(row)
    if isinstance(d.get("standard_info"), str):
        d["standard_info"] = json.loads(d["standard_info"])
    return d


async def _get_object_or_404(
    db: asyncpg.Connection,
    project_id: str,
    object_id: str,
) -> asyncpg.Record:
    row = await db.fetchrow(
        "SELECT * FROM ppg_project_objects WHERE id = $1 AND project_id = $2",
        object_id, project_id,
    )
    if not row:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Object not found"})
    return row


# ---------------------------------------------------------------------------
# CRUD — Project Objects
# ---------------------------------------------------------------------------

@router.get("/{project_id}/objects")
async def list_objects(
    user: CurrentUser,
    project_id: UUID,
    object_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    q: Optional[str] = Query(None, description="Search by name or code"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    conditions: list[str] = ["project_id = $1"]
    params: list = [str(project_id)]
    idx = 2

    if object_type:
        conditions.append(f"object_type = ${idx}")
        params.append(object_type)
        idx += 1
    if status:
        conditions.append(f"status = ${idx}")
        params.append(status)
        idx += 1
    if q:
        conditions.append(f"(name ILIKE ${idx} OR code ILIKE ${idx})")
        params.append(f"%{q}%")
        idx += 1

    where = " AND ".join(conditions)
    total_row = await db.fetchrow(
        f"SELECT COUNT(*) as cnt FROM ppg_project_objects WHERE {where}", *params
    )
    total = total_row["cnt"] if total_row else 0

    offset = (page - 1) * size
    rows = await db.fetch(
        f"SELECT * FROM ppg_project_objects WHERE {where} "
        f"ORDER BY created_at DESC LIMIT ${idx} OFFSET ${idx+1}",
        *params, size, offset,
    )
    return {
        "data": [_parse_obj(r) for r in rows],
        "meta": {"total": total, "page": page, "size": size},
    }


@router.post("/{project_id}/objects", status_code=201)
async def create_object(
    user: CurrentUser,
    project_id: UUID,
    body: ProjectObjectCreate,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    obj_id = str(uuid4())
    try:
        row = await db.fetchrow(
            """
            INSERT INTO ppg_project_objects
                (id, project_id, object_type, name, code, description, owner, standard_info, created_by)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
            RETURNING *
            """,
            obj_id,
            str(project_id),
            body.object_type,
            body.name,
            body.code,
            body.description,
            body.owner,
            body.standard_info,
            user.sub,
        )
    except asyncpg.UniqueViolationError:
        raise HTTPException(
            409,
            detail={
                "code": "CONFLICT",
                "message": f"Code '{body.code}' already exists in this project",
            },
        )
    await log_audit(
        db=db,
        entity_type="ppg_project_objects",
        entity_id=obj_id,
        action="CREATE",
        changed_by=user.sub,
        new_values=body.model_dump(),
    )
    return {"data": _parse_obj(row)}


@router.get("/{project_id}/objects/export")
async def export_objects(
    user: CurrentUser,
    project_id: UUID,
    object_type: str = Query(..., description="web_app | mobile_app | api | elt"),
    db: asyncpg.Connection = Depends(get_db),
) -> StreamingResponse:
    """Export objects to Excel — BR-012 fixed template per type."""
    if object_type not in OBJECT_TYPES:
        raise HTTPException(422, detail={"code": "VALIDATION_ERROR", "message": "Invalid object_type"})

    rows = await db.fetch(
        "SELECT * FROM ppg_project_objects WHERE project_id = $1 AND object_type = $2 "
        "AND status != 'decommissioned' ORDER BY name",
        str(project_id), object_type,
    )

    # Column templates per type — BR-012 fixed, no customization allowed
    columns_map: dict[str, list[str]] = {
        "web_app": [
            "name", "code", "description", "owner", "status",
            "tech_stack", "version", "url_dev", "url_staging", "url_uat", "url_prod",
            "deployment_type", "notes",
        ],
        "mobile_app": [
            "name", "code", "description", "owner", "status",
            "platform", "version", "store_link_ios", "store_link_android",
            "tech_stack", "min_os_version", "notes",
        ],
        "api": [
            "name", "code", "description", "owner", "status",
            "base_url", "auth_method", "version", "protocol",
            "url_dev", "url_uat", "swagger_url", "notes",
        ],
        "elt": [
            "name", "code", "description", "owner", "status",
            "source_system", "target_system", "schedule", "technology",
            "data_format", "volume_estimate", "sla_minutes", "notes",
        ],
    }
    columns = columns_map[object_type]

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            500,
            detail={"code": "INTERNAL_ERROR", "message": "openpyxl not installed"},
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = object_type
    ws.append(columns)

    for row in rows:
        d = _parse_obj(row)
        info = d.get("standard_info", {})
        excel_row: list = []
        for col in columns:
            if col in ("name", "code", "description", "owner", "status"):
                excel_row.append(d.get(col) or "")
            else:
                excel_row.append(info.get(col) or "")
        ws.append(excel_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date

    filename = f"{object_type}_export_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/{project_id}/objects/import")
async def import_objects(
    user: CurrentUser,
    project_id: UUID,
    object_type: str = Form(...),
    conflict_strategy: str = Form("ask"),
    file: UploadFile = File(...),
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    """Import objects from Excel — BR-013 conflict handling."""
    if object_type not in OBJECT_TYPES:
        raise HTTPException(422, detail={"code": "VALIDATION_ERROR", "message": "Invalid object_type"})
    if conflict_strategy not in ("ask", "overwrite", "skip"):
        raise HTTPException(422, detail={"code": "VALIDATION_ERROR", "message": "conflict_strategy must be ask|overwrite|skip"})

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            500,
            detail={"code": "INTERNAL_ERROR", "message": "openpyxl not installed"},
        )

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        return {"data": {"created": 0, "updated": 0, "skipped": 0, "errors": []}}

    headers = [str(h).strip() if h else "" for h in rows_iter[0]]

    # Check which codes already exist (BR-013)
    existing_codes = {
        row["code"]
        for row in await db.fetch(
            "SELECT code FROM ppg_project_objects WHERE project_id = $1", str(project_id)
        )
    }

    type_map: dict = {
        "web_app": WebAppInfo,
        "mobile_app": MobileAppInfo,
        "api": ApiInfo,
        "elt": EltInfo,
    }
    model_cls = type_map[object_type]

    conflicting_codes: list[str] = []
    data_rows: list[dict] = []
    errors: list[dict] = []

    for i, raw_row in enumerate(rows_iter[1:], start=2):
        row_dict: dict = {headers[j]: (raw_row[j] if j < len(raw_row) else None) for j in range(len(headers))}
        code = str(row_dict.get("code", "")).strip().upper()
        if not code:
            errors.append({"row": i, "error": "code is required"})
            continue
        if code in existing_codes:
            conflicting_codes.append(code)
        data_rows.append({"code": code, "row": i, "data": row_dict})

    if conflicting_codes and conflict_strategy == "ask":
        raise HTTPException(
            409,
            detail={
                "code": "IMPORT_CONFLICT",
                "message": "Found duplicate objects. Please confirm action.",
                "details": {
                    "conflicting_codes": conflicting_codes,
                    "hint": "Resend with conflict_strategy=overwrite or conflict_strategy=skip",
                },
            },
        )

    created = 0
    updated = 0
    skipped = 0

    for item in data_rows:
        code = item["code"]
        row_dict = item["data"]
        row_num = item["row"]

        if code in existing_codes:
            if conflict_strategy == "skip":
                skipped += 1
                continue
            # overwrite path handled below

        # Build standard_info from type model fields
        info_fields = {
            f: row_dict.get(f)
            for f in model_cls.model_fields
            if f != "object_type"
        }
        # Validate
        try:
            validated = model_cls.model_validate({"object_type": object_type, **info_fields})
        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})
            continue

        name = str(row_dict.get("name", "")).strip()
        description = row_dict.get("description")
        owner = row_dict.get("owner")
        status = row_dict.get("status", "active")

        if code in existing_codes and conflict_strategy == "overwrite":
            # Upsert — update
            await db.execute(
                """
                UPDATE ppg_project_objects
                SET name=$3, description=$4, owner=$5, status=$6, standard_info=$7,
                    updated_at=NOW(), updated_by=$8
                WHERE project_id=$1 AND code=$2
                """,
                str(project_id), code, name, description, owner, status,
                validated.model_dump(exclude={"object_type"}),
                user.sub if user else "import",
            )
            updated += 1
        else:
            try:
                obj_id = str(uuid4())
                await db.execute(
                    """
                    INSERT INTO ppg_project_objects
                        (id, project_id, object_type, name, code, description, owner, status,
                         standard_info, created_by)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)
                    """,
                    obj_id, str(project_id), object_type, name, code,
                    description, owner, status,
                    validated.model_dump(exclude={"object_type"}),
                    user.sub if user else "import",
                )
                created += 1
            except asyncpg.UniqueViolationError:
                errors.append({"row": row_num, "error": f"code '{code}' conflict"})

    return {"data": {"created": created, "updated": updated, "skipped": skipped, "errors": errors}}


@router.get("/{project_id}/objects/{object_id}")
async def get_object(
    user: CurrentUser,
    project_id: UUID,
    object_id: UUID,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    row = await _get_object_or_404(db, str(project_id), str(object_id))
    obj = _parse_obj(row)

    # Attach stats
    ba_count = await db.fetchval(
        """
        SELECT COUNT(*) FROM ba_document_object_links
        WHERE object_id = $1
        """,
        str(object_id),
    )
    tc_count = await db.fetchval(
        "SELECT COUNT(*) FROM test_case_object_links WHERE object_id = $1",
        str(object_id),
    )
    obj["stats"] = {
        "ba_docs_linked": ba_count or 0,
        "test_cases_linked": tc_count or 0,
    }
    return {"data": obj}


@router.put("/{project_id}/objects/{object_id}")
async def update_object(
    user: CurrentUser,
    project_id: UUID,
    object_id: UUID,
    body: ProjectObjectUpdate,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    """Update object — object_type and code are immutable after creation."""
    existing = await _get_object_or_404(db, str(project_id), str(object_id))
    old_values = _parse_obj(existing)

    updates: dict = {}
    if body.name is not None:
        updates["name"] = body.name
    if body.description is not None:
        updates["description"] = body.description
    if body.owner is not None:
        updates["owner"] = body.owner
    if body.standard_info is not None:
        # Re-validate standard_info against existing object_type
        otype = existing["object_type"]
        type_map: dict = {
            "web_app": WebAppInfo,
            "mobile_app": MobileAppInfo,
            "api": ApiInfo,
            "elt": EltInfo,
        }
        model_cls = type_map[otype]
        model_cls.model_validate({"object_type": otype, **body.standard_info})
        updates["standard_info"] = body.standard_info

    if not updates:
        raise HTTPException(400, detail={"code": "VALIDATION_ERROR", "message": "No fields to update"})

    set_parts = [f"{k} = ${i + 3}" for i, k in enumerate(updates.keys())]
    row = await db.fetchrow(
        f"UPDATE ppg_project_objects SET {', '.join(set_parts)}, updated_at = NOW(), updated_by = $2 "
        f"WHERE id = $1 RETURNING *",
        str(object_id), user.sub, *updates.values(),
    )
    await log_audit(
        db=db,
        entity_type="ppg_project_objects",
        entity_id=str(object_id),
        action="UPDATE",
        changed_by=user.sub,
        old_values=old_values,
        new_values=body.model_dump(exclude_none=True),
    )
    return {"data": _parse_obj(row)}


@router.delete("/{project_id}/objects/{object_id}", status_code=204, response_class=Response)
async def decommission_object(
    user: CurrentUser,
    project_id: UUID,
    object_id: UUID,
    db: asyncpg.Connection = Depends(get_db),
) -> Response:
    """Soft delete — set status = decommissioned. Block if active BA docs or test cases exist."""
    await _get_object_or_404(db, str(project_id), str(object_id))

    # Check active BA docs (table ba_document_object_links created by schema-draft-v2 V009)
    # Gracefully handle pre-migration state when table doesn't exist yet
    try:
        active_ba = await db.fetchval(
            """
            SELECT COUNT(*) FROM ba_document_object_links dol
            JOIN ba_documents d ON d.id = dol.document_id
            WHERE dol.object_id = $1 AND d.status NOT IN ('archived')
            """,
            str(object_id),
        )
    except Exception:
        active_ba = 0
    if active_ba and active_ba > 0:
        raise HTTPException(
            409,
            detail={
                "code": "CONFLICT",
                "message": "Cannot decommission: object has active BA documents linked",
            },
        )

    result = await db.execute(
        "UPDATE ppg_project_objects SET status = 'decommissioned', updated_at = NOW(), updated_by = $2 "
        "WHERE id = $1",
        str(object_id), user.sub,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Object not found"})

    await log_audit(
        db=db,
        entity_type="ppg_project_objects",
        entity_id=str(object_id),
        action="DELETE",
        changed_by=user.sub,
    )
    return Response(status_code=204)


# ---------------------------------------------------------------------------
# Connections
# ---------------------------------------------------------------------------

@router.get("/{project_id}/objects/{object_id}/connections")
async def list_connections(
    user: CurrentUser,
    project_id: UUID,
    object_id: UUID,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    await _get_object_or_404(db, str(project_id), str(object_id))
    oid = str(object_id)

    out_rows = await db.fetch(
        """
        SELECT c.*, po.name as target_name, po.object_type as target_type,
               po.project_id as target_project_id
        FROM ppg_object_connections c
        JOIN ppg_project_objects po ON po.id = c.target_object_id
        WHERE c.source_object_id = $1 AND c.status != 'removed'
        ORDER BY c.created_at DESC
        """,
        oid,
    )
    in_rows = await db.fetch(
        """
        SELECT c.*, po.name as source_name, po.object_type as source_type,
               po.project_id as source_project_id
        FROM ppg_object_connections c
        JOIN ppg_project_objects po ON po.id = c.source_object_id
        WHERE c.target_object_id = $1 AND c.status != 'removed'
        ORDER BY c.created_at DESC
        """,
        oid,
    )

    return {
        "data": {
            "outbound": [dict(r) for r in out_rows],
            "inbound": [dict(r) for r in in_rows],
        }
    }


@router.post("/{project_id}/objects/{object_id}/connections", status_code=201)
async def create_connection(
    user: CurrentUser,
    project_id: UUID,
    object_id: UUID,
    body: ConnectionCreate,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    await _get_object_or_404(db, str(project_id), str(object_id))

    # Validate target exists
    target = await db.fetchrow(
        "SELECT id FROM ppg_project_objects WHERE id = $1", str(body.target_object_id)
    )
    if not target:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Target object not found"})

    if str(object_id) == str(body.target_object_id):
        raise HTTPException(
            409,
            detail={"code": "CONFLICT", "message": "Self-loop connections are not allowed"},
        )

    conn_id = str(uuid4())
    try:
        row = await db.fetchrow(
            """
            INSERT INTO ppg_object_connections
                (id, source_object_id, target_object_id, connection_type, protocol,
                 frequency, description, created_by)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8)
            RETURNING *
            """,
            conn_id,
            str(object_id),
            str(body.target_object_id),
            body.connection_type,
            body.protocol,
            body.frequency,
            body.description,
            user.sub,
        )
    except asyncpg.UniqueViolationError:
        raise HTTPException(
            409,
            detail={
                "code": "CONFLICT",
                "message": "Connection of this type already exists between these objects",
            },
        )

    await log_audit(
        db=db,
        entity_type="ppg_object_connections",
        entity_id=conn_id,
        action="CREATE",
        changed_by=user.sub,
        new_values=body.model_dump(),
    )
    return {"data": dict(row)}


@router.delete("/{project_id}/objects/{object_id}/connections/{conn_id}", status_code=204, response_class=Response)
async def remove_connection(
    user: CurrentUser,
    project_id: UUID,
    object_id: UUID,
    conn_id: UUID,
    db: asyncpg.Connection = Depends(get_db),
) -> Response:
    result = await db.execute(
        "UPDATE ppg_object_connections SET status = 'removed', updated_at = NOW(), updated_by = $2 "
        "WHERE id = $1 AND source_object_id = $3",
        str(conn_id), user.sub, str(object_id),
    )
    if result == "UPDATE 0":
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Connection not found"})

    await log_audit(
        db=db,
        entity_type="ppg_object_connections",
        entity_id=str(conn_id),
        action="DELETE",
        changed_by=user.sub,
    )
    return Response(status_code=204)
