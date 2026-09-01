"""
Request Management Router — v7
POST/GET/PATCH/DELETE /requests/change-requests   (change_requests)
POST/GET/PATCH/DELETE /requests/service           (service_requests)
"""
import io
from datetime import datetime, date
import os
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File as UpFile
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
import asyncpg
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.auth import CurrentUser
from app.database import get_db

router = APIRouter(prefix="/requests", tags=["requests"])

_default_uploads = Path(__file__).resolve().parent.parent.parent.parent / "uploads"
UPLOAD_DIR = Path(os.getenv("UPLOADS_DIR", str(_default_uploads))).resolve()

# ── Constants ─────────────────────────────────────────────────────────────────

CR_CHANGE_TYPES = {"scope", "timeline", "resource", "budget", "technical", "process", "other"}
CR_STATUSES     = {"submitted", "reviewing", "approved", "rejected", "implementing", "implemented", "cancelled"}
SR_REQUEST_TYPES = {"bug_fix", "enhancement", "support", "incident", "access_request", "data_request", "other"}
SR_STATUSES      = {"submitted", "reviewing", "approved", "in_progress", "resolved", "rejected", "cancelled"}
CR_KINDS         = {"standard", "internal"}
PRIORITIES       = {"critical", "high", "medium", "low"}
SEVERITIES       = {"critical", "high", "medium", "low"}
ENVIRONMENTS     = {"DEV", "SIT", "UAT", "PROD", "DR", "STAGING"}

# ── Schemas ───────────────────────────────────────────────────────────────────

class CRCreate(BaseModel):
    # QUYỀN SỞ HỮU — sản phẩm mà CR thay đổi. Bắt buộc với CR tạo mới (V052).
    product_id:    str  = Field(..., min_length=1)
    # QUY KẾT NGUỒN — dự án tài trợ, tùy chọn. Trống với CR phát sinh sau khi dự án đóng.
    project_id:    Optional[str]  = None
    # standard = CR nghiệp vụ (phải qua BRS + test) | internal = CR nội bộ sửa tay Master Doc
    cr_kind:       str  = Field("standard")
    title:         str  = Field(..., min_length=1, max_length=255)
    description:   Optional[str]  = None
    change_type:   str  = Field("other")
    priority:      str  = Field("medium")
    impact_scope:  Optional[str]  = None
    impact_effort: Optional[str]  = None
    requested_by:  str  = Field(..., min_length=1)
    assigned_to:   Optional[str]  = None
    target_date:   Optional[str]  = None
    notes:         Optional[str]  = None


class CRPatch(BaseModel):
    title:         Optional[str]  = Field(None, min_length=1, max_length=255)
    product_id:    Optional[str]  = None
    # Cho phép gắn/đổi dự án tài trợ sau khi tạo — CR sinh ngoài dự án rồi được quy kết về
    # một dự án là chuyện bình thường. cr_kind KHÔNG cho sửa: đổi CR nghiệp vụ thành nội bộ
    # sẽ làm mất bắt buộc BRS/test của một thay đổi đã ghi nhận.
    project_id:    Optional[str]  = None
    description:   Optional[str]  = None
    change_type:   Optional[str]  = None
    priority:      Optional[str]  = None
    status:        Optional[str]  = None
    impact_scope:  Optional[str]  = None
    impact_effort: Optional[str]  = None
    assigned_to:   Optional[str]  = None
    target_date:   Optional[str]  = None
    approved_by:   Optional[str]  = None
    notes:         Optional[str]  = None
    comment:       Optional[str]  = None


class SRCreate(BaseModel):
    product_id:   Optional[str]  = None
    title:        str  = Field(..., min_length=1, max_length=255)
    description:  Optional[str]  = None
    request_type: str  = Field("support")
    priority:     str  = Field("medium")
    severity:     Optional[str]  = None
    environment:  Optional[str]  = None
    requested_by: str  = Field(..., min_length=1)
    assigned_to:  Optional[str]  = None
    sla_deadline: Optional[str]  = None


class SRPatch(BaseModel):
    title:            Optional[str]  = Field(None, min_length=1, max_length=255)
    description:      Optional[str]  = None
    request_type:     Optional[str]  = None
    priority:         Optional[str]  = None
    severity:         Optional[str]  = None
    environment:      Optional[str]  = None
    status:           Optional[str]  = None
    assigned_to:      Optional[str]  = None
    sla_deadline:     Optional[str]  = None
    resolution_notes: Optional[str]  = None
    comment:          Optional[str]  = None


# ── Helpers ───────────────────────────────────────────────────────────────────

async def _next_cr_code(db: asyncpg.Connection) -> str:
    seq = await db.fetchval("SELECT nextval('cr_seq')")
    return f"CR-{datetime.now().year}-{seq:03d}"


async def _next_sr_code(db: asyncpg.Connection) -> str:
    seq = await db.fetchval("SELECT nextval('sr_seq')")
    return f"SR-{datetime.now().year}-{seq:03d}"


def _validate_cr_create(body: CRCreate) -> None:
    if body.change_type not in CR_CHANGE_TYPES:
        raise HTTPException(400, f"change_type không hợp lệ: {body.change_type}")
    if body.priority not in PRIORITIES:
        raise HTTPException(400, f"priority không hợp lệ: {body.priority}")
    if body.cr_kind not in CR_KINDS:
        raise HTTPException(400, f"cr_kind không hợp lệ: {body.cr_kind}")


def _validate_cr_patch(body: CRPatch) -> None:
    if body.change_type and body.change_type not in CR_CHANGE_TYPES:
        raise HTTPException(400, f"change_type không hợp lệ: {body.change_type}")
    if body.priority and body.priority not in PRIORITIES:
        raise HTTPException(400, f"priority không hợp lệ: {body.priority}")
    if body.status and body.status not in CR_STATUSES:
        raise HTTPException(400, f"status không hợp lệ: {body.status}")


def _validate_sr_create(body: SRCreate) -> None:
    if body.request_type not in SR_REQUEST_TYPES:
        raise HTTPException(400, f"request_type không hợp lệ: {body.request_type}")
    if body.priority not in PRIORITIES:
        raise HTTPException(400, f"priority không hợp lệ: {body.priority}")
    if body.severity and body.severity not in SEVERITIES:
        raise HTTPException(400, f"severity không hợp lệ: {body.severity}")
    if body.environment and body.environment not in ENVIRONMENTS:
        raise HTTPException(400, f"environment không hợp lệ: {body.environment}")


def _validate_sr_patch(body: SRPatch) -> None:
    if body.request_type and body.request_type not in SR_REQUEST_TYPES:
        raise HTTPException(400, f"request_type không hợp lệ: {body.request_type}")
    if body.priority and body.priority not in PRIORITIES:
        raise HTTPException(400, f"priority không hợp lệ: {body.priority}")
    if body.severity and body.severity not in SEVERITIES:
        raise HTTPException(400, f"severity không hợp lệ: {body.severity}")
    if body.environment and body.environment not in ENVIRONMENTS:
        raise HTTPException(400, f"environment không hợp lệ: {body.environment}")
    if body.status and body.status not in SR_STATUSES:
        raise HTTPException(400, f"status không hợp lệ: {body.status}")


async def _log_history(
    db: asyncpg.Connection,
    ref_type: str,
    ref_id: str,
    action: str,
    actor: str,
    from_status: Optional[str] = None,
    to_status: Optional[str] = None,
    comment: Optional[str] = None,
) -> None:
    await db.execute(
        """
        INSERT INTO request_history (ref_type, ref_id, action, actor, from_status, to_status, comment)
        VALUES ($1, $2::uuid, $3, $4, $5, $6, $7)
        """,
        ref_type, ref_id, action, actor, from_status, to_status, comment,
    )


# ── Project Change Requests ───────────────────────────────────────────────────

@router.get("/change-requests")
async def list_project_changes(
    user: CurrentUser,
    project_id:  Optional[str] = Query(None),
    product_id:  Optional[str] = Query(None),
    status:      Optional[str] = Query(None),
    priority:    Optional[str] = Query(None),
    change_type: Optional[str] = Query(None),
    db: asyncpg.Connection = Depends(get_db),
):
    clauses: list[str] = []
    params:  list      = []
    i = 1

    if project_id:
        clauses.append(f"cr.project_id = ${i}::uuid"); params.append(project_id); i += 1
    if product_id:
        clauses.append(f"cr.product_id = ${i}::uuid"); params.append(product_id); i += 1
    if status:
        clauses.append(f"cr.status = ${i}");     params.append(status);     i += 1
    if priority:
        clauses.append(f"cr.priority = ${i}");   params.append(priority);   i += 1
    if change_type:
        clauses.append(f"cr.change_type = ${i}"); params.append(change_type); i += 1

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    rows = await db.fetch(
        f"""
        SELECT cr.*, p.name AS project_name, p.code AS project_code,
               cp.product_name, cp.product_code
        FROM change_requests cr
        LEFT JOIN projects p ON p.id = cr.project_id
        LEFT JOIN catalog_products cp ON cp.id = cr.product_id
        {where}
        ORDER BY cr.created_at DESC
        """,
        *params,
    )
    return [dict(r) for r in rows]


@router.post("/change-requests", status_code=201)
async def create_project_change(
    user: CurrentUser,
    body: CRCreate,
    db: asyncpg.Connection = Depends(get_db),
):
    _validate_cr_create(body)

    # Quyền sở hữu: sản phẩm bắt buộc tồn tại (V052)
    product = await db.fetchval(
        "SELECT id FROM catalog_products WHERE id=$1::uuid", body.product_id
    )
    if not product:
        raise HTTPException(404, f"Sản phẩm '{body.product_id}' không tồn tại")

    # Quy kết nguồn: dự án tùy chọn, nhưng nếu có thì phải tồn tại
    if body.project_id:
        exists = await db.fetchval(
            "SELECT id FROM projects WHERE id=$1::uuid", body.project_id
        )
        if not exists:
            raise HTTPException(404, f"Project '{body.project_id}' không tồn tại")

    code = await _next_cr_code(db)
    target = date.fromisoformat(body.target_date) if body.target_date else None

    row = await db.fetchrow(
        """
        INSERT INTO change_requests
            (request_code, project_id, product_id, cr_kind, title, description,
             change_type, priority, impact_scope, impact_effort, requested_by,
             assigned_to, target_date, notes)
        VALUES ($1,$2::uuid,$3::uuid,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14)
        RETURNING *
        """,
        code, body.project_id, body.product_id, body.cr_kind, body.title, body.description,
        body.change_type, body.priority,
        body.impact_scope, body.impact_effort,
        body.requested_by, body.assigned_to, target, body.notes,
    )
    await _log_history(db, 'cr', str(row['id']), 'created', body.requested_by,
                       to_status='submitted')
    return dict(row)


@router.get("/change-requests/export")
async def export_project_changes(
    user: CurrentUser,
    project_id:  Optional[str] = Query(None),
    status:      Optional[str] = Query(None),
    priority:    Optional[str] = Query(None),
    change_type: Optional[str] = Query(None),
    db: asyncpg.Connection = Depends(get_db),
):
    clauses: list[str] = []
    params:  list      = []
    i = 1
    if project_id:
        clauses.append(f"cr.project_id = ${i}::uuid"); params.append(project_id); i += 1
    if status:
        clauses.append(f"cr.status = ${i}");     params.append(status);     i += 1
    if priority:
        clauses.append(f"cr.priority = ${i}");   params.append(priority);   i += 1
    if change_type:
        clauses.append(f"cr.change_type = ${i}"); params.append(change_type); i += 1

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    rows = await db.fetch(
        f"""
        SELECT cr.*, p.name AS project_name, p.code AS project_code
        FROM change_requests cr
        LEFT JOIN projects p ON p.id = cr.project_id
        {where}
        ORDER BY cr.created_at DESC
        """,
        *params,
    )
    if not rows:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CR"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=cr_export.xlsx"})

    ids = [str(r['id']) for r in rows]
    ph  = ",".join(f"${j+1}::uuid" for j in range(len(ids)))
    hist_rows = await db.fetch(
        f"SELECT * FROM request_history WHERE ref_type='cr' AND ref_id IN ({ph}) ORDER BY ref_id, created_at ASC",
        *ids,
    )
    hist_map: dict[str, list] = {}
    for h in hist_rows:
        hist_map.setdefault(str(h['ref_id']), []).append(h)

    STATUS_VI = {
        'submitted': 'Khởi tạo', 'reviewing': 'Đang review', 'approved': 'Đã duyệt',
        'implementing': 'Đang triển khai', 'implemented': 'Đã triển khai',
        'rejected': 'Từ chối', 'cancelled': 'Hủy',
    }
    ACTION_VI = {'created': 'Khởi tạo', 'status_changed': 'Chuyển trạng thái', 'updated': 'Cập nhật'}
    CHANGE_VI = {
        'scope': 'Phạm vi', 'timeline': 'Thời gian', 'resource': 'Nguồn lực',
        'budget': 'Ngân sách', 'technical': 'Kỹ thuật', 'process': 'Quy trình', 'other': 'Khác',
    }
    PRIORITY_VI = {'critical': 'Khẩn cấp', 'high': 'Cao', 'medium': 'Trung bình', 'low': 'Thấp'}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CR"

    headers = ["Code", "Tiêu đề", "Dự án", "Loại thay đổi", "Ưu tiên", "Trạng thái",
               "Người yêu cầu", "Người xử lý", "Ngày mục tiêu", "Ngày tạo", "Lịch sử"]
    header_fill = PatternFill("solid", fgColor="003366")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, r in enumerate(rows, 2):
        rid = str(r['id'])
        hist_lines = []
        for h in hist_map.get(rid, []):
            ts = h['created_at'].strftime('%d/%m/%Y %H:%M') if h['created_at'] else ''
            action_label = ACTION_VI.get(h['action'], h['action'])
            transition = ""
            if h['from_status'] or h['to_status']:
                frm = STATUS_VI.get(h['from_status'], h['from_status'] or '') if h['from_status'] else ''
                to  = STATUS_VI.get(h['to_status'],   h['to_status']   or '') if h['to_status']   else ''
                transition = f": {frm} → {to}" if frm or to else ""
            comment_part = f" | {h['comment']}" if h['comment'] else ""
            hist_lines.append(f"{ts} | {h['actor']} | {action_label}{transition}{comment_part}")
        hist_text = "\n".join(hist_lines)

        ws.cell(row=row_idx, column=1,  value=r['request_code'])
        ws.cell(row=row_idx, column=2,  value=r['title'])
        ws.cell(row=row_idx, column=3,  value=r.get('project_name') or '')
        ws.cell(row=row_idx, column=4,  value=CHANGE_VI.get(r['change_type'], r['change_type']))
        ws.cell(row=row_idx, column=5,  value=PRIORITY_VI.get(r['priority'], r['priority']))
        ws.cell(row=row_idx, column=6,  value=STATUS_VI.get(r['status'], r['status']))
        ws.cell(row=row_idx, column=7,  value=r['requested_by'])
        ws.cell(row=row_idx, column=8,  value=r.get('assigned_to') or '')
        ws.cell(row=row_idx, column=9,  value=str(r['target_date']) if r.get('target_date') else '')
        ws.cell(row=row_idx, column=10, value=r['created_at'].strftime('%d/%m/%Y %H:%M') if r.get('created_at') else '')
        hist_cell = ws.cell(row=row_idx, column=11, value=hist_text)
        hist_cell.alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = [14, 40, 24, 16, 12, 18, 18, 18, 14, 18, 60]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=cr_export.xlsx"})


@router.get("/change-requests/{cr_id}")
async def get_project_change(
    user: CurrentUser,
    cr_id: str,
    db: asyncpg.Connection = Depends(get_db),
):
    row = await db.fetchrow(
        """
        SELECT cr.*,
               p.name  AS project_name, p.code AS project_code,
               cp.product_name, cp.product_code, cp.product_type, cp.domain_code
        FROM change_requests cr
        LEFT JOIN projects p           ON p.id  = cr.project_id
        LEFT JOIN catalog_products cp  ON cp.id = cr.product_id
        WHERE cr.id = $1::uuid
        """,
        cr_id,
    )
    if not row:
        raise HTTPException(404, "CR không tồn tại")
    return dict(row)


@router.patch("/change-requests/{cr_id}")
async def update_project_change(
    user: CurrentUser,
    cr_id: str,
    body: CRPatch,
    db: asyncpg.Connection = Depends(get_db),
):
    _validate_cr_patch(body)
    updates = body.model_dump(exclude_none=True)
    comment = updates.pop('comment', None)
    if not updates:
        raise HTTPException(400, "Không có trường nào để cập nhật")

    # Fetch current status for history
    current = await db.fetchrow(
        "SELECT status FROM change_requests WHERE id=$1::uuid", cr_id
    )
    if not current:
        raise HTTPException(404, "CR không tồn tại")
    old_status = current['status']

    # Kiểm tra tham chiếu tồn tại TRƯỚC khi UPDATE. Không có bước này thì UUID sai sẽ
    # vi phạm khóa ngoại và trả 500 thô, người dùng không hiểu vì sao.
    if updates.get("product_id"):
        ok = await db.fetchval(
            "SELECT 1 FROM catalog_products WHERE id=$1::uuid", updates["product_id"]
        )
        if not ok:
            raise HTTPException(404, f"Sản phẩm '{updates['product_id']}' không tồn tại")
    if updates.get("project_id"):
        ok = await db.fetchval(
            "SELECT 1 FROM projects WHERE id=$1::uuid", updates["project_id"]
        )
        if not ok:
            raise HTTPException(404, f"Dự án '{updates['project_id']}' không tồn tại")

    # Auto set approved_at
    if updates.get("status") == "approved" and "approved_by" not in updates:
        updates.setdefault("approved_by", user.sub)

    sets: list[str] = []
    params: list    = []
    i = 1
    for key, val in updates.items():
        if key == "target_date" and val:
            val = date.fromisoformat(val)
        if key == "status" and val == "approved":
            sets.append(f"approved_at = NOW()")
        cast = "::uuid" if key in ("product_id", "project_id") else ""
        sets.append(f"{key} = ${i}{cast}"); params.append(val); i += 1
    sets.append("updated_at = NOW()")

    result = await db.execute(
        f"UPDATE change_requests SET {', '.join(sets)} WHERE id = ${i}::uuid",
        *params, cr_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "CR không tồn tại")

    new_status = updates.get('status')
    action = 'status_changed' if new_status and new_status != old_status else 'updated'
    await _log_history(db, 'cr', cr_id, action, user.sub,
                       from_status=old_status if new_status else None,
                       to_status=new_status or None,
                       comment=comment)
    return {"ok": True}


@router.delete("/change-requests/{cr_id}", status_code=204)
async def delete_project_change(
    user: CurrentUser,
    cr_id: str,
    db: asyncpg.Connection = Depends(get_db),
):
    result = await db.execute(
        "DELETE FROM change_requests WHERE id = $1::uuid", cr_id
    )
    if result == "DELETE 0":
        raise HTTPException(404, "CR không tồn tại")


@router.get("/change-requests/{cr_id}/history")
async def get_cr_history(
    user: CurrentUser,
    cr_id: str,
    db: asyncpg.Connection = Depends(get_db),
):
    rows = await db.fetch(
        """
        SELECT * FROM request_history
        WHERE ref_type = 'cr' AND ref_id = $1::uuid
        ORDER BY created_at ASC
        """,
        cr_id,
    )
    return [dict(r) for r in rows]


# ── Service Requests ──────────────────────────────────────────────────────────

@router.get("/service")
async def list_service_requests(
    user: CurrentUser,
    product_id:   Optional[str] = Query(None),
    status:       Optional[str] = Query(None),
    priority:     Optional[str] = Query(None),
    request_type: Optional[str] = Query(None),
    environment:  Optional[str] = Query(None),
    db: asyncpg.Connection = Depends(get_db),
):
    clauses: list[str] = []
    params:  list      = []
    i = 1

    if product_id:
        clauses.append(f"sr.product_id = ${i}::uuid"); params.append(product_id); i += 1
    if status:
        clauses.append(f"sr.status = ${i}");           params.append(status);     i += 1
    if priority:
        clauses.append(f"sr.priority = ${i}");         params.append(priority);   i += 1
    if request_type:
        clauses.append(f"sr.request_type = ${i}");     params.append(request_type); i += 1
    if environment:
        clauses.append(f"sr.environment = ${i}");      params.append(environment); i += 1

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    rows = await db.fetch(
        f"""
        SELECT sr.*, cp.product_name, cp.product_type
        FROM service_requests sr
        LEFT JOIN catalog_products cp ON cp.id = sr.product_id
        {where}
        ORDER BY sr.created_at DESC
        """,
        *params,
    )
    return [dict(r) for r in rows]


@router.post("/service", status_code=201)
async def create_service_request(
    user: CurrentUser,
    body: SRCreate,
    db: asyncpg.Connection = Depends(get_db),
):
    _validate_sr_create(body)

    if body.product_id:
        exists = await db.fetchval(
            "SELECT id FROM catalog_products WHERE id=$1::uuid", body.product_id
        )
        if not exists:
            raise HTTPException(404, f"Product '{body.product_id}' không tồn tại")

    code = await _next_sr_code(db)
    sla  = datetime.fromisoformat(body.sla_deadline) if body.sla_deadline else None

    row = await db.fetchrow(
        """
        INSERT INTO service_requests
            (request_code, product_id, title, description, request_type, priority,
             severity, environment, requested_by, assigned_to, sla_deadline)
        VALUES ($1,$2::uuid,$3,$4,$5,$6,$7,$8,$9,$10,$11)
        RETURNING *
        """,
        code, body.product_id, body.title, body.description,
        body.request_type, body.priority,
        body.severity, body.environment,
        body.requested_by, body.assigned_to, sla,
    )
    await _log_history(db, 'sr', str(row['id']), 'created', body.requested_by,
                       to_status='submitted')
    return dict(row)


@router.get("/service/export")
async def export_service_requests(
    user: CurrentUser,
    product_id:   Optional[str] = Query(None),
    status:       Optional[str] = Query(None),
    priority:     Optional[str] = Query(None),
    request_type: Optional[str] = Query(None),
    environment:  Optional[str] = Query(None),
    db: asyncpg.Connection = Depends(get_db),
):
    clauses: list[str] = []
    params:  list      = []
    i = 1
    if product_id:
        clauses.append(f"sr.product_id = ${i}::uuid"); params.append(product_id); i += 1
    if status:
        clauses.append(f"sr.status = ${i}");           params.append(status);     i += 1
    if priority:
        clauses.append(f"sr.priority = ${i}");         params.append(priority);   i += 1
    if request_type:
        clauses.append(f"sr.request_type = ${i}");     params.append(request_type); i += 1
    if environment:
        clauses.append(f"sr.environment = ${i}");      params.append(environment); i += 1

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    rows = await db.fetch(
        f"""
        SELECT sr.*, cp.product_name, cp.product_type
        FROM service_requests sr
        LEFT JOIN catalog_products cp ON cp.id = sr.product_id
        {where}
        ORDER BY sr.created_at DESC
        """,
        *params,
    )
    if not rows:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SR"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=sr_export.xlsx"})

    ids = [str(r['id']) for r in rows]
    ph  = ",".join(f"${j+1}::uuid" for j in range(len(ids)))
    hist_rows = await db.fetch(
        f"SELECT * FROM request_history WHERE ref_type='sr' AND ref_id IN ({ph}) ORDER BY ref_id, created_at ASC",
        *ids,
    )
    hist_map: dict[str, list] = {}
    for h in hist_rows:
        hist_map.setdefault(str(h['ref_id']), []).append(h)

    STATUS_VI = {
        'submitted': 'Khởi tạo', 'reviewing': 'Đang review', 'approved': 'Đã duyệt',
        'in_progress': 'Đang xử lý', 'resolved': 'Đã xử lý',
        'rejected': 'Từ chối', 'cancelled': 'Hủy',
    }
    ACTION_VI = {'created': 'Khởi tạo', 'status_changed': 'Chuyển trạng thái', 'updated': 'Cập nhật'}
    TYPE_VI = {
        'bug_fix': 'Sửa lỗi', 'enhancement': 'Cải tiến', 'support': 'Hỗ trợ',
        'incident': 'Sự cố', 'access_request': 'Cấp quyền', 'data_request': 'Yêu cầu dữ liệu', 'other': 'Khác',
    }
    PRIORITY_VI = {'critical': 'Khẩn cấp', 'high': 'Cao', 'medium': 'Trung bình', 'low': 'Thấp'}
    SEVERITY_VI = {'critical': 'Nghiêm trọng', 'high': 'Cao', 'medium': 'Trung bình', 'low': 'Thấp'}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SR"

    headers = ["Code", "Tiêu đề", "Ứng dụng/Sản phẩm", "Loại yêu cầu", "Ưu tiên",
               "Độ nghiêm trọng", "Môi trường", "Trạng thái",
               "Người yêu cầu", "Người xử lý", "SLA Deadline", "Ngày tạo", "Lịch sử"]
    header_fill = PatternFill("solid", fgColor="003366")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, r in enumerate(rows, 2):
        rid = str(r['id'])
        hist_lines = []
        for h in hist_map.get(rid, []):
            ts = h['created_at'].strftime('%d/%m/%Y %H:%M') if h['created_at'] else ''
            action_label = ACTION_VI.get(h['action'], h['action'])
            transition = ""
            if h['from_status'] or h['to_status']:
                frm = STATUS_VI.get(h['from_status'], h['from_status'] or '') if h['from_status'] else ''
                to  = STATUS_VI.get(h['to_status'],   h['to_status']   or '') if h['to_status']   else ''
                transition = f": {frm} → {to}" if frm or to else ""
            comment_part = f" | {h['comment']}" if h['comment'] else ""
            hist_lines.append(f"{ts} | {h['actor']} | {action_label}{transition}{comment_part}")
        hist_text = "\n".join(hist_lines)

        ws.cell(row=row_idx, column=1,  value=r['request_code'])
        ws.cell(row=row_idx, column=2,  value=r['title'])
        ws.cell(row=row_idx, column=3,  value=r.get('product_name') or '')
        ws.cell(row=row_idx, column=4,  value=TYPE_VI.get(r['request_type'], r['request_type']))
        ws.cell(row=row_idx, column=5,  value=PRIORITY_VI.get(r['priority'], r['priority']))
        ws.cell(row=row_idx, column=6,  value=SEVERITY_VI.get(r['severity'], r['severity']) if r.get('severity') else '')
        ws.cell(row=row_idx, column=7,  value=r.get('environment') or '')
        ws.cell(row=row_idx, column=8,  value=STATUS_VI.get(r['status'], r['status']))
        ws.cell(row=row_idx, column=9,  value=r['requested_by'])
        ws.cell(row=row_idx, column=10, value=r.get('assigned_to') or '')
        ws.cell(row=row_idx, column=11, value=str(r['sla_deadline']) if r.get('sla_deadline') else '')
        ws.cell(row=row_idx, column=12, value=r['created_at'].strftime('%d/%m/%Y %H:%M') if r.get('created_at') else '')
        hist_cell = ws.cell(row=row_idx, column=13, value=hist_text)
        hist_cell.alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = [14, 40, 24, 16, 12, 16, 12, 18, 18, 18, 18, 18, 60]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=sr_export.xlsx"})


@router.get("/service/{sr_id}")
async def get_service_request(
    user: CurrentUser,
    sr_id: str,
    db: asyncpg.Connection = Depends(get_db),
):
    row = await db.fetchrow(
        """
        SELECT sr.*, cp.product_name, cp.product_type
        FROM service_requests sr
        LEFT JOIN catalog_products cp ON cp.id = sr.product_id
        WHERE sr.id = $1::uuid
        """,
        sr_id,
    )
    if not row:
        raise HTTPException(404, "Service Request không tồn tại")
    return dict(row)


@router.patch("/service/{sr_id}")
async def update_service_request(
    user: CurrentUser,
    sr_id: str,
    body: SRPatch,
    db: asyncpg.Connection = Depends(get_db),
):
    _validate_sr_patch(body)
    updates = body.model_dump(exclude_none=True)
    comment = updates.pop('comment', None)
    if not updates:
        raise HTTPException(400, "Không có trường nào để cập nhật")

    current = await db.fetchrow(
        "SELECT status FROM service_requests WHERE id=$1::uuid", sr_id
    )
    if not current:
        raise HTTPException(404, "Service Request không tồn tại")
    old_status = current['status']

    sets: list[str] = []
    params: list    = []
    i = 1
    for key, val in updates.items():
        if key == "sla_deadline" and val:
            val = datetime.fromisoformat(val)
        sets.append(f"{key} = ${i}"); params.append(val); i += 1

    if updates.get("status") in ("resolved", "closed"):
        sets.append("resolved_at = NOW()")
    sets.append("updated_at = NOW()")

    result = await db.execute(
        f"UPDATE service_requests SET {', '.join(sets)} WHERE id = ${i}::uuid",
        *params, sr_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Service Request không tồn tại")

    new_status = updates.get('status')
    action = 'status_changed' if new_status and new_status != old_status else 'updated'
    await _log_history(db, 'sr', sr_id, action, user.sub,
                       from_status=old_status if new_status else None,
                       to_status=new_status or None,
                       comment=comment)
    return {"ok": True}


@router.delete("/service/{sr_id}", status_code=204)
async def delete_service_request(
    user: CurrentUser,
    sr_id: str,
    db: asyncpg.Connection = Depends(get_db),
):
    result = await db.execute(
        "DELETE FROM service_requests WHERE id = $1::uuid", sr_id
    )
    if result == "DELETE 0":
        raise HTTPException(404, "Service Request không tồn tại")


@router.get("/service/{sr_id}/history")
async def get_sr_history(
    user: CurrentUser,
    sr_id: str,
    db: asyncpg.Connection = Depends(get_db),
):
    rows = await db.fetch(
        """
        SELECT * FROM request_history
        WHERE ref_type = 'sr' AND ref_id = $1::uuid
        ORDER BY created_at ASC
        """,
        sr_id,
    )
    return [dict(r) for r in rows]


# ── Attachments ───────────────────────────────────────────────────────────────

def _safe_filename(name: str, dest_dir: Path) -> tuple[str, Path]:
    """Return (safe_name, dest_path) — appends _N suffix if file exists."""
    safe = Path(name).name or "file"
    dest = dest_dir / safe
    if dest.exists():
        stem, suffix, i = Path(safe).stem, Path(safe).suffix, 1
        while dest.exists():
            dest = dest_dir / f"{stem}_{i}{suffix}"; i += 1
        safe = dest.name
    return safe, dest


async def _save_attachment(
    db: asyncpg.Connection,
    ref_type: str,
    ref_id: str,
    file: UploadFile,
    actor: str,
) -> dict:
    dest_dir = UPLOAD_DIR / "requests" / ref_type / ref_id
    dest_dir.mkdir(parents=True, exist_ok=True)
    safe_name, dest = _safe_filename(file.filename or "file", dest_dir)
    contents = await file.read()
    dest.write_bytes(contents)
    row = await db.fetchrow(
        """
        INSERT INTO request_attachments
            (ref_type, ref_id, filename, file_size, mime_type, stored_path, uploaded_by)
        VALUES ($1, $2::uuid, $3, $4, $5, $6, $7)
        RETURNING *
        """,
        ref_type, ref_id, safe_name, len(contents),
        file.content_type, str(dest), actor,
    )
    return dict(row)


@router.post("/change-requests/{cr_id}/attachments", status_code=201)
async def upload_cr_attachment(
    user: CurrentUser,
    cr_id: str,
    file: UploadFile = UpFile(...),
    db: asyncpg.Connection = Depends(get_db),
):
    if not await db.fetchval("SELECT id FROM change_requests WHERE id=$1::uuid", cr_id):
        raise HTTPException(404, "CR không tồn tại")
    return await _save_attachment(db, "cr", cr_id, file, user.sub)


@router.get("/change-requests/{cr_id}/attachments")
async def list_cr_attachments(
    user: CurrentUser,
    cr_id: str,
    db: asyncpg.Connection = Depends(get_db),
):
    rows = await db.fetch(
        "SELECT * FROM request_attachments WHERE ref_type='cr' AND ref_id=$1::uuid ORDER BY created_at",
        cr_id,
    )
    return [dict(r) for r in rows]


@router.post("/service/{sr_id}/attachments", status_code=201)
async def upload_sr_attachment(
    user: CurrentUser,
    sr_id: str,
    file: UploadFile = UpFile(...),
    db: asyncpg.Connection = Depends(get_db),
):
    if not await db.fetchval("SELECT id FROM service_requests WHERE id=$1::uuid", sr_id):
        raise HTTPException(404, "Service Request không tồn tại")
    return await _save_attachment(db, "sr", sr_id, file, user.sub)


@router.get("/service/{sr_id}/attachments")
async def list_sr_attachments(
    user: CurrentUser,
    sr_id: str,
    db: asyncpg.Connection = Depends(get_db),
):
    rows = await db.fetch(
        "SELECT * FROM request_attachments WHERE ref_type='sr' AND ref_id=$1::uuid ORDER BY created_at",
        sr_id,
    )
    return [dict(r) for r in rows]


@router.get("/attachments/{att_id}/download")
async def download_attachment(
    user: CurrentUser,
    att_id: str,
    db: asyncpg.Connection = Depends(get_db),
):
    row = await db.fetchrow("SELECT * FROM request_attachments WHERE id=$1::uuid", att_id)
    if not row:
        raise HTTPException(404, "File không tồn tại")
    path = Path(row["stored_path"])
    if not path.exists():
        raise HTTPException(404, "File đã bị xóa khỏi hệ thống")
    return FileResponse(str(path), filename=row["filename"])
