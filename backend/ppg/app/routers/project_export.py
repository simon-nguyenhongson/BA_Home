"""
Project Export / Import Router
GET  /projects/{project_id}/export  — Download XLSX (4 sheets)
POST /projects/{project_id}/import  — Upload XLSX, upsert data
"""
import io
import json
import logging
from datetime import date, datetime
from uuid import uuid4
from typing import Optional

import asyncpg
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from app.auth import CurrentUser
from app.database import get_db

router = APIRouter(prefix="/projects", tags=["project-export"])
logger = logging.getLogger(__name__)

# ── Style helpers ─────────────────────────────────────────────────────────────

_H1_FILL   = PatternFill("solid", fgColor="1E3A5F")   # dark navy
_H2_FILL   = PatternFill("solid", fgColor="2E6DA4")   # mid blue
_H3_FILL   = PatternFill("solid", fgColor="D9E8F5")   # light blue
_ALT_FILL  = PatternFill("solid", fgColor="F4F8FC")   # stripe

_H1_FONT   = Font(bold=True, color="FFFFFF", size=11)
_H2_FONT   = Font(bold=True, color="FFFFFF", size=10)
_H3_FONT   = Font(bold=True, color="1E3A5F", size=9)
_BODY_FONT = Font(size=9)

_THIN = Side(style="thin", color="C5D5E4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header(ws, row: int, col: int, value: str, level: int = 2, colspan: int = 1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font  = [_H1_FONT, _H2_FONT, _H3_FONT][level - 1]
    cell.fill  = [_H1_FILL, _H2_FILL, _H3_FILL][level - 1]
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _BORDER
    if colspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + colspan - 1
        )
    return cell


def _cell(ws, row: int, col: int, value, wrap: bool = False, bold: bool = False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font   = Font(size=9, bold=bold)
    cell.border = _BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    return cell


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze(ws, cell: str = "A2"):
    ws.freeze_panes = cell


def _fmt_date(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]


def _arr(v) -> str:
    """Convert list / JSON string → newline-joined string."""
    if v is None:
        return ""
    if isinstance(v, list):
        return "\n".join(str(x) for x in v)
    if isinstance(v, str):
        try:
            parsed = json.loads(v)
            if isinstance(parsed, list):
                return "\n".join(str(x) for x in parsed)
        except Exception:
            pass
    return str(v)


# ── Data fetch helpers ────────────────────────────────────────────────────────

async def _fetch_all(db, project_id: str) -> dict:
    project = await db.fetchrow("SELECT * FROM projects WHERE id=$1", project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    brief = await db.fetchrow(
        "SELECT * FROM project_briefs WHERE project_id=$1", project_id
    )
    milestones = await db.fetch(
        "SELECT * FROM project_milestones WHERE project_id=$1 ORDER BY track, sort_order",
        project_id,
    )
    members = await db.fetch(
        "SELECT * FROM project_members WHERE project_id=$1 ORDER BY role, full_name",
        project_id,
    )
    stakeholders = await db.fetch(
        "SELECT * FROM project_stakeholders WHERE project_id=$1 ORDER BY name",
        project_id,
    )
    ba_tasks = await db.fetch(
        """SELECT bt.*, pm.name AS milestone_name
           FROM ba_tasks bt
           LEFT JOIN project_milestones pm ON pm.id = bt.milestone_id
           WHERE bt.project_id=$1
           ORDER BY pm.sort_order, bt.title""",
        project_id,
    )
    test_tasks = await db.fetch(
        """SELECT tt.*, pm.name AS milestone_name
           FROM test_tasks tt
           LEFT JOIN project_milestones pm ON pm.id = tt.milestone_id
           WHERE tt.project_id=$1
           ORDER BY pm.sort_order, tt.title""",
        project_id,
    )
    stage_gates = await db.fetch(
        "SELECT * FROM project_stage_gates WHERE project_id=$1 ORDER BY stage_order",
        project_id,
    )
    return dict(
        project=dict(project),
        brief=dict(brief) if brief else {},
        milestones=[dict(m) for m in milestones],
        members=[dict(m) for m in members],
        stakeholders=[dict(s) for s in stakeholders],
        ba_tasks=[dict(t) for t in ba_tasks],
        test_tasks=[dict(t) for t in test_tasks],
        stage_gates=[dict(g) for g in stage_gates],
    )


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_overview(wb, data: dict):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False
    p = data["project"]
    b = data["brief"]

    r = 1
    _header(ws, r, 1, "PROJECT OVERVIEW", level=1, colspan=2); r += 1
    ws.row_dimensions[1].height = 22

    def kv(label, value, wrap=False):
        nonlocal r
        _cell(ws, r, 1, label, bold=True)
        _cell(ws, r, 2, value, wrap=wrap)
        if wrap:
            ws.row_dimensions[r].height = max(30, min(120, 15 * (str(value or "").count("\n") + 1)))
        r += 1

    # ── Basic info
    _header(ws, r, 1, "🗂 Thông tin dự án", level=2, colspan=2); r += 1
    kv("Mã dự án",   p.get("code"))
    kv("Tên dự án",  p.get("name"))
    kv("Trạng thái", p.get("status"))
    kv("Owner",      p.get("owner"))
    kv("Ngày bắt đầu", _fmt_date(p.get("start_date")))
    kv("Ngày kết thúc", _fmt_date(p.get("end_date")))
    kv("Mô tả",      p.get("description"), wrap=True)

    # ── Brief sections
    if b:
        _header(ws, r, 1, "🎯 Mục tiêu & Tổng quan", level=2, colspan=2); r += 1
        kv("Mục đích",           b.get("purpose"), wrap=True)
        kv("Thông tin chung",    b.get("general_info"), wrap=True)
        kv("Success Metrics",    _arr(b.get("success_metrics")), wrap=True)
        kv("Giá trị end-user",   b.get("enduser_value"), wrap=True)

        _header(ws, r, 1, "👥 Đối tượng người dùng", level=2, colspan=2); r += 1
        kv("Primary Users",  b.get("primary_users"), wrap=True)
        kv("Pain Points",    b.get("pain_points"), wrap=True)
        kv("User Role Matrix", _arr(b.get("user_role_matrix")), wrap=True)

        _header(ws, r, 1, "📋 Yêu cầu chức năng", level=2, colspan=2); r += 1
        kv("Must Have",        _arr(b.get("must_have_features")), wrap=True)
        kv("Nice to Have",     _arr(b.get("nice_to_have_features")), wrap=True)
        kv("System Integrations", _arr(b.get("system_integrations")), wrap=True)

        _header(ws, r, 1, "⚙️ Yêu cầu phi chức năng", level=2, colspan=2); r += 1
        kv("Performance/Scalability", b.get("performance_scalability"), wrap=True)
        kv("Compliance/Security",     b.get("compliance_security"), wrap=True)
        kv("Availability/Reliability", b.get("availability_reliability"), wrap=True)

        _header(ws, r, 1, "📊 Dữ liệu & Báo cáo", level=2, colspan=2); r += 1
        kv("Nhu cầu dữ liệu",  b.get("data_needs"), wrap=True)
        kv("Nhu cầu báo cáo",  b.get("reporting_needs"), wrap=True)

        _header(ws, r, 1, "⚠️ Ràng buộc & Rủi ro", level=2, colspan=2); r += 1
        kv("Ràng buộc thời gian", b.get("time_constraints"), wrap=True)
        kv("Dependencies",        _arr(b.get("dependencies")), wrap=True)
        kv("Rủi ro tiềm năng",    _arr(b.get("potential_risks")), wrap=True)

        _header(ws, r, 1, "🗓 Timeline & Phương pháp", level=2, colspan=2); r += 1
        kv("Methodology",         b.get("methodology"))
        kv("Decision Makers",     _arr(b.get("decision_makers")), wrap=True)
        kv("Key Milestone Notes", _arr(b.get("key_milestones_notes")), wrap=True)

    _set_col_widths(ws, [28, 80])


def _build_timeline(wb, data: dict):
    ws = wb.create_sheet("Timeline")
    ws.sheet_view.showGridLines = False

    COLS = ["Track", "Tên Milestone", "Loại", "Mô tả", "Ngày bắt đầu", "Ngày kết thúc",
            "Trạng thái", "Done Criteria", "Sort"]
    _header(ws, 1, 1, "PROJECT TIMELINE — MILESTONES", level=1, colspan=len(COLS))
    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(COLS, 1):
        _header(ws, 2, ci, h, level=3)
    _freeze(ws, "A3")

    track_order = {"project": "1", "ba": "2", "test": "3"}
    for ri, m in enumerate(sorted(data["milestones"],
                                   key=lambda x: (track_order.get(x.get("track", "project"), "9"),
                                                  x.get("sort_order", 0))), 3):
        fill = _ALT_FILL if ri % 2 == 0 else None
        track_label = {"project": "🔵 Project", "ba": "📄 BA", "test": "🧪 Test"}.get(
            m.get("track", "project"), m.get("track", "")
        )
        vals = [
            track_label,
            m.get("name"),
            m.get("milestone_type"),
            m.get("description"),
            _fmt_date(m.get("start_date")),
            _fmt_date(m.get("end_date")),
            m.get("status"),
            m.get("done_criteria"),
            m.get("sort_order"),
        ]
        for ci, v in enumerate(vals, 1):
            c = _cell(ws, ri, ci, v, wrap=(ci in (4, 8)))
            if fill:
                c.fill = fill

    _set_col_widths(ws, [12, 30, 22, 40, 14, 14, 12, 40, 6])


def _build_resources(wb, data: dict):
    ws = wb.create_sheet("Nguồn lực")
    ws.sheet_view.showGridLines = False
    r = 1

    # ── Team Members ──────────────────────────────────────────
    _header(ws, r, 1, "TEAM MEMBERS", level=1, colspan=5); r += 1
    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(["Họ tên", "Vai trò", "Email", "Alias", "Trạng thái"], 1):
        _header(ws, r, ci, h, level=3)
    r += 1
    _freeze(ws, "A3")

    for i, m in enumerate(data["members"]):
        fill = _ALT_FILL if i % 2 == 0 else None
        vals = [m.get("full_name"), m.get("role"), m.get("email"),
                m.get("alias"), "Active" if m.get("is_active") else "Inactive"]
        for ci, v in enumerate(vals, 1):
            c = _cell(ws, r, ci, v)
            if fill:
                c.fill = fill
        r += 1

    if not data["members"]:
        _cell(ws, r, 1, "(Chưa có thành viên)")
        r += 1

    r += 1  # blank separator

    # ── Stakeholders ──────────────────────────────────────────
    _header(ws, r, 1, "STAKEHOLDERS", level=1, colspan=6); r += 1
    for ci, h in enumerate(["Tên", "Vai trò", "Tổ chức", "Interest", "Influence", "Engagement Strategy"], 1):
        _header(ws, r, ci, h, level=3)
    r += 1

    for i, s in enumerate(data["stakeholders"]):
        fill = _ALT_FILL if i % 2 == 0 else None
        vals = [s.get("name"), s.get("role"), s.get("organization"),
                s.get("interest_level"), s.get("influence_level"),
                s.get("engagement_strategy")]
        for ci, v in enumerate(vals, 1):
            c = _cell(ws, r, ci, v, wrap=(ci == 6))
            if fill:
                c.fill = fill
        r += 1

    if not data["stakeholders"]:
        _cell(ws, r, 1, "(Chưa có stakeholder)")

    _set_col_widths(ws, [25, 16, 20, 12, 12, 40])


def _build_todo(wb, data: dict):
    ws = wb.create_sheet("To-do list")
    ws.sheet_view.showGridLines = False
    r = 1

    # ── BA Tasks ──────────────────────────────────────────────
    _header(ws, r, 1, "BA TASKS", level=1, colspan=5); r += 1
    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(["Tiêu đề", "Loại task", "Milestone", "Due Date", "Trạng thái"], 1):
        _header(ws, r, ci, h, level=3)
    r += 1
    _freeze(ws, "A3")

    for i, t in enumerate(data["ba_tasks"]):
        fill = _ALT_FILL if i % 2 == 0 else None
        vals = [t.get("title"), t.get("task_type"),
                t.get("milestone_name"), _fmt_date(t.get("due_date")), t.get("status")]
        for ci, v in enumerate(vals, 1):
            c = _cell(ws, r, ci, v)
            if fill:
                c.fill = fill
        r += 1

    if not data["ba_tasks"]:
        _cell(ws, r, 1, "(Chưa có BA task)")
        r += 1

    r += 1

    # ── Test Tasks ────────────────────────────────────────────
    _header(ws, r, 1, "TEST TASKS", level=1, colspan=5); r += 1
    for ci, h in enumerate(["Tiêu đề", "Loại task", "Milestone", "Due Date", "Trạng thái"], 1):
        _header(ws, r, ci, h, level=3)
    r += 1

    for i, t in enumerate(data["test_tasks"]):
        fill = _ALT_FILL if i % 2 == 0 else None
        vals = [t.get("title"), t.get("task_type"),
                t.get("milestone_name"), _fmt_date(t.get("due_date")), t.get("status")]
        for ci, v in enumerate(vals, 1):
            c = _cell(ws, r, ci, v)
            if fill:
                c.fill = fill
        r += 1

    if not data["test_tasks"]:
        _cell(ws, r, 1, "(Chưa có Test task)")
        r += 1

    r += 1

    # ── Stage Gates ───────────────────────────────────────────
    _header(ws, r, 1, "STAGE GATES", level=1, colspan=6); r += 1
    for ci, h in enumerate(["Stage Name", "Order", "Status", "Gate Date", "Sign Off By", "Notes"], 1):
        _header(ws, r, ci, h, level=3)
    r += 1

    for i, g in enumerate(data["stage_gates"]):
        fill = _ALT_FILL if i % 2 == 0 else None
        vals = [g.get("stage_name"), g.get("stage_order"), g.get("status"),
                _fmt_date(g.get("gate_date")), g.get("sign_off_by"), g.get("notes")]
        for ci, v in enumerate(vals, 1):
            c = _cell(ws, r, ci, v, wrap=(ci == 6))
            if fill:
                c.fill = fill
        r += 1

    if not data["stage_gates"]:
        _cell(ws, r, 1, "(Chưa có stage gate)")

    _set_col_widths(ws, [40, 7, 14, 12, 20, 50])


# ── Export endpoint ───────────────────────────────────────────────────────────

@router.get("/{project_id}/export")
async def export_project(
    user: CurrentUser,
    project_id: str,
    db: asyncpg.Connection = Depends(get_db),
):
    """Export toàn bộ thông tin project ra file XLSX (4 sheets)."""
    data = await _fetch_all(db, project_id)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _build_overview(wb, data)
    _build_timeline(wb, data)
    _build_resources(wb, data)
    _build_todo(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    project_code = data["project"].get("code", project_id)
    today = date.today().strftime("%Y%m%d")
    filename = f"project_{project_code}_{today}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Import endpoint ───────────────────────────────────────────────────────────

def _read_kv_sheet(ws) -> dict[str, str]:
    """Read a 2-column key→value sheet (col A = label, col B = value)."""
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip():
            key = str(row[0]).strip()
            val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            result[key] = val
    return result


def _parse_arr(val: str) -> list[str]:
    if not val or not str(val).strip():
        return []
    return [x.strip() for x in str(val).split("\n") if x.strip()]


@router.post("/{project_id}/import", status_code=200)
async def import_project(
    user: CurrentUser,
    project_id: str,
    file: UploadFile = File(...),
    db: asyncpg.Connection = Depends(get_db),
):
    """
    Import thông tin project từ XLSX (4 sheets).
    - Overview  → upsert project info + brief
    - Timeline  → update milestone status / dates (match by name+track, no delete)
    - Nguồn lực → upsert members (match by full_name+role) + stakeholders (match by name)
    - To-do list → update ba_tasks / test_tasks / stage_gates status (match by title/stage_name)
    """
    # Validate project exists
    row = await db.fetchrow("SELECT id, code FROM projects WHERE id=$1", project_id)
    if not row:
        raise HTTPException(404, "Project not found")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Không đọc được file XLSX: {e}")

    result: dict = {"updated": {}, "errors": []}

    # ── Sheet 1: Overview ─────────────────────────────────────
    if "Overview" in wb.sheetnames:
        try:
            ws = wb["Overview"]
            brief_map: dict[str, str] = {}
            project_map: dict[str, str] = {}

            BASIC_KEYS = {
                "Tên dự án": "name", "Trạng thái": "status",
                "Owner": "owner", "Mô tả": "description",
            }
            BRIEF_MAP = {
                "Mục đích": "purpose", "Thông tin chung": "general_info",
                "Success Metrics": "success_metrics", "Giá trị end-user": "enduser_value",
                "Primary Users": "primary_users", "Pain Points": "pain_points",
                "User Role Matrix": "user_role_matrix",
                "Must Have": "must_have_features", "Nice to Have": "nice_to_have_features",
                "System Integrations": "system_integrations",
                "Performance/Scalability": "performance_scalability",
                "Compliance/Security": "compliance_security",
                "Availability/Reliability": "availability_reliability",
                "Nhu cầu dữ liệu": "data_needs", "Nhu cầu báo cáo": "reporting_needs",
                "Ràng buộc thời gian": "time_constraints",
                "Dependencies": "dependencies", "Rủi ro tiềm năng": "potential_risks",
                "Methodology": "methodology",
                "Decision Makers": "decision_makers",
                "Key Milestone Notes": "key_milestones_notes",
            }
            ARR_FIELDS = {
                "success_metrics", "user_role_matrix", "must_have_features",
                "nice_to_have_features", "system_integrations", "dependencies",
                "potential_risks", "decision_makers", "key_milestones_notes",
            }

            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                label = str(row_cells[0]).strip() if row_cells[0] else ""
                val   = str(row_cells[1]).strip() if len(row_cells) > 1 and row_cells[1] is not None else ""
                if label in BASIC_KEYS:
                    project_map[BASIC_KEYS[label]] = val
                if label in BRIEF_MAP:
                    field = BRIEF_MAP[label]
                    brief_map[field] = _parse_arr(val) if field in ARR_FIELDS else val

            # Update project basic fields
            if project_map:
                set_parts = [f"{k}=${i+2}" for i, k in enumerate(project_map.keys())]
                await db.execute(
                    f"UPDATE projects SET {', '.join(set_parts)}, updated_at=NOW() WHERE id=$1",
                    project_id, *project_map.values(),
                )

            # Upsert brief
            if brief_map:
                cols   = list(brief_map.keys())
                vals_b = [brief_map[c] for c in cols]
                set_parts = [f"{c}=${i+3}" for i, c in enumerate(cols)]
                await db.execute(
                    f"""INSERT INTO project_briefs (id, project_id, {', '.join(cols)})
                        VALUES ($1, $2, {', '.join(f'${i+3}' for i in range(len(cols)))})
                        ON CONFLICT (project_id)
                        DO UPDATE SET {', '.join(set_parts)}, updated_at=NOW()""",
                    str(uuid4()), project_id, *vals_b,
                )

            result["updated"]["overview"] = True
        except Exception as e:
            result["errors"].append(f"Overview: {e}")

    # ── Sheet 2: Timeline ─────────────────────────────────────
    if "Timeline" in wb.sheetnames:
        try:
            ws = wb["Timeline"]
            updated_ms = 0
            for row_cells in ws.iter_rows(min_row=3, values_only=True):
                if not row_cells[1]:
                    continue
                track_raw = str(row_cells[0] or "").replace("🔵 ", "").replace("📄 ", "").replace("🧪 ", "").strip().lower()
                track_map = {"project": "project", "ba": "ba", "test": "test"}
                track = track_map.get(track_raw, "project")
                name    = str(row_cells[1]).strip()
                start   = str(row_cells[4]).strip() if row_cells[4] else None
                end     = str(row_cells[5]).strip() if row_cells[5] else None
                status  = str(row_cells[6]).strip() if row_cells[6] else None
                done_c  = str(row_cells[7]).strip() if row_cells[7] else None

                updates: dict = {}
                if status:  updates["status"]         = status
                if start:   updates["start_date"]     = start
                if end:     updates["end_date"]       = end
                if done_c:  updates["done_criteria"]  = done_c

                if updates:
                    set_parts = [f"{k}=${i+3}" for i, k in enumerate(updates.keys())]
                    await db.execute(
                        f"UPDATE project_milestones SET {', '.join(set_parts)}, updated_at=NOW() "
                        f"WHERE project_id=$1 AND track=$2 AND name=$3",
                        project_id, track, name, *updates.values(),
                    )
                    updated_ms += 1

            result["updated"]["timeline"] = updated_ms
        except Exception as e:
            result["errors"].append(f"Timeline: {e}")

    # ── Sheet 3: Nguồn lực ────────────────────────────────────
    if "Nguồn lực" in wb.sheetnames:
        ws = wb["Nguồn lực"]
        rows = list(ws.iter_rows(values_only=True))

        # Find section separators
        member_rows: list = []
        stakeholder_rows: list = []
        current_section: Optional[str] = None

        for r_vals in rows:
            first = str(r_vals[0] or "").strip().upper()
            if "TEAM MEMBERS" in first:
                current_section = "members"
                continue
            if "STAKEHOLDERS" in first:
                current_section = "stakeholders"
                continue
            if not any(r_vals):
                continue
            if r_vals[0] in ("Họ tên", "Tên"):  # header rows
                continue
            if current_section == "members" and r_vals[0]:
                member_rows.append(r_vals)
            elif current_section == "stakeholders" and r_vals[0]:
                stakeholder_rows.append(r_vals)

        updated_m = 0
        try:
            for mr in member_rows:
                full_name = str(mr[0]).strip() if mr[0] else None
                role      = str(mr[1]).strip() if len(mr) > 1 and mr[1] else None
                email     = str(mr[2]).strip() if len(mr) > 2 and mr[2] else None
                alias     = str(mr[3]).strip() if len(mr) > 3 and mr[3] else None
                is_active = str(mr[4] or "").lower() != "inactive" if len(mr) > 4 else True

                if not full_name:
                    continue
                existing = await db.fetchrow(
                    "SELECT id FROM project_members WHERE project_id=$1 AND full_name=$2",
                    project_id, full_name,
                )
                if existing:
                    await db.execute(
                        "UPDATE project_members SET role=$3, email=$4, alias=$5, is_active=$6 WHERE id=$1 AND project_id=$2",
                        existing["id"], project_id, role, email, alias, is_active,
                    )
                else:
                    await db.execute(
                        "INSERT INTO project_members (id, project_id, full_name, role, email, alias, is_active) "
                        "VALUES ($1,$2,$3,$4,$5,$6,$7)",
                        str(uuid4()), project_id, full_name, role, email, alias, is_active,
                    )
                updated_m += 1
            result["updated"]["members"] = updated_m
        except Exception as e:
            result["errors"].append(f"Members: {e}")

        updated_s = 0
        try:
            for sr in stakeholder_rows:
                name     = str(sr[0]).strip() if sr[0] else None
                role     = str(sr[1]).strip() if len(sr) > 1 and sr[1] else None
                org      = str(sr[2]).strip() if len(sr) > 2 and sr[2] else None
                interest = str(sr[3]).strip() if len(sr) > 3 and sr[3] else None
                influence= str(sr[4]).strip() if len(sr) > 4 and sr[4] else None
                strategy = str(sr[5]).strip() if len(sr) > 5 and sr[5] else None

                if not name:
                    continue
                existing = await db.fetchrow(
                    "SELECT id FROM project_stakeholders WHERE project_id=$1 AND name=$2",
                    project_id, name,
                )
                if existing:
                    await db.execute(
                        "UPDATE project_stakeholders SET role=$3,organization=$4,interest_level=$5,influence_level=$6,engagement_strategy=$7 "
                        "WHERE id=$1 AND project_id=$2",
                        existing["id"], project_id, role, org, interest, influence, strategy,
                    )
                else:
                    await db.execute(
                        "INSERT INTO project_stakeholders (id,project_id,name,role,organization,interest_level,influence_level,engagement_strategy) "
                        "VALUES ($1,$2,$3,$4,$5,$6,$7,$8)",
                        str(uuid4()), project_id, name, role, org,
                        interest or "medium", influence or "medium", strategy,
                    )
                updated_s += 1
            result["updated"]["stakeholders"] = updated_s
        except Exception as e:
            result["errors"].append(f"Stakeholders: {e}")

    # ── Sheet 4: To-do list ───────────────────────────────────
    if "To-do list" in wb.sheetnames:
        ws = wb["To-do list"]
        rows = list(ws.iter_rows(values_only=True))

        ba_rows, test_rows, gate_rows = [], [], []
        current_section = None

        for r_vals in rows:
            first = str(r_vals[0] or "").strip().upper()
            if "BA TASKS" in first:
                current_section = "ba"
                continue
            if "TEST TASKS" in first:
                current_section = "test"
                continue
            if "STAGE GATES" in first:
                current_section = "gates"
                continue
            if not any(r_vals):
                continue
            if r_vals[0] in ("Tiêu đề", "Stage Name"):
                continue
            if current_section == "ba" and r_vals[0]:
                ba_rows.append(r_vals)
            elif current_section == "test" and r_vals[0]:
                test_rows.append(r_vals)
            elif current_section == "gates" and r_vals[0]:
                gate_rows.append(r_vals)

        try:
            updated_bt = 0
            for t in ba_rows:
                title  = str(t[0]).strip() if t[0] else None
                status = str(t[4]).strip() if len(t) > 4 and t[4] else None
                if title and status:
                    await db.execute(
                        "UPDATE ba_tasks SET status=$3 WHERE project_id=$1 AND title=$2",
                        project_id, title, status,
                    )
                    updated_bt += 1
            result["updated"]["ba_tasks"] = updated_bt
        except Exception as e:
            result["errors"].append(f"BA tasks: {e}")

        try:
            updated_tt = 0
            for t in test_rows:
                title  = str(t[0]).strip() if t[0] else None
                status = str(t[4]).strip() if len(t) > 4 and t[4] else None
                if title and status:
                    await db.execute(
                        "UPDATE test_tasks SET status=$3 WHERE project_id=$1 AND title=$2",
                        project_id, title, status,
                    )
                    updated_tt += 1
            result["updated"]["test_tasks"] = updated_tt
        except Exception as e:
            result["errors"].append(f"Test tasks: {e}")

        try:
            updated_g = 0
            for g in gate_rows:
                stage_name = str(g[0]).strip() if g[0] else None
                status     = str(g[2]).strip() if len(g) > 2 and g[2] else None
                gate_date  = str(g[3]).strip() if len(g) > 3 and g[3] else None
                sign_off   = str(g[4]).strip() if len(g) > 4 and g[4] else None
                notes      = str(g[5]).strip() if len(g) > 5 and g[5] else None

                if not stage_name:
                    continue
                updates: dict = {}
                if status:    updates["status"]      = status
                if gate_date: updates["gate_date"]   = gate_date
                if sign_off:  updates["sign_off_by"] = sign_off
                if notes:     updates["notes"]       = notes

                if updates:
                    set_parts = [f"{k}=${i+3}" for i, k in enumerate(updates.keys())]
                    await db.execute(
                        f"UPDATE project_stage_gates SET {', '.join(set_parts)}, updated_at=NOW() "
                        f"WHERE project_id=$1 AND stage_name=$2",
                        project_id, stage_name, *updates.values(),
                    )
                    updated_g += 1
            result["updated"]["stage_gates"] = updated_g
        except Exception as e:
            result["errors"].append(f"Stage gates: {e}")

    return result
