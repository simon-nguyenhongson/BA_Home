"""
Automation Test Router — task test theo CR, gắn với BRS đã duyệt.

Luồng (docs/design/AI-DOC-AUTOMATION-FLOW.md mục 3):
  BRS approved → task 'need_test' (tạo tự động ở cr_brs.py)
  → [Gen testcase] (AI skill gen_test_case từ BRS)
  → map test case với test case đã record trong Capture Studio
  → chạy nhiều lần, import kết quả run
  → [Gen report] (AI skill gen_test_report)
  → [Close] → [Export] XLSX
"""
from __future__ import annotations

import io
import json
import re
from datetime import datetime
from typing import Optional
from uuid import uuid4

import asyncpg
from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

from app.auth import CurrentUser
from app.database import get_db
from app.services.ai_agent import assert_skill_for_step, run_skill
from app.services.ai_sse import sse_response, step_done, step_running, vi_num
from app.services.audit_service import log_audit

router = APIRouter(prefix="/automation", tags=["automation"])

VALID_PRIORITIES = {"critical", "high", "medium", "low"}


class GenerateCasesRequest(BaseModel):
    skill_code: str = "gen_test_case"
    note: str = ""


class TestCaseCreate(BaseModel):
    code: str = Field(..., max_length=40)
    title: str = Field(..., max_length=300)
    precondition: str = ""
    steps: str = ""
    expected: str = ""
    priority: str = "medium"


class TestCaseUpdate(BaseModel):
    code: Optional[str] = Field(None, max_length=40)
    title: Optional[str] = Field(None, max_length=300)
    precondition: Optional[str] = None
    steps: Optional[str] = None
    expected: Optional[str] = None
    priority: Optional[str] = None
    studio_tc_id: Optional[str] = None
    status: Optional[str] = None
    sort_order: Optional[int] = None


class RunImport(BaseModel):
    run_ref: str = ""
    summary: dict = Field(default_factory=dict)


class ReportRequest(BaseModel):
    skill_code: str = "gen_test_report"


async def _get_task_or_404(db: asyncpg.Connection, task_id: str) -> asyncpg.Record:
    row = await db.fetchrow(
        """
        SELECT t.*, cr.request_code, cr.title AS cr_title, cr.description AS cr_description,
               cr.project_id, p.code AS project_code, p.name AS project_name,
               b.title AS brs_title, b.content AS brs_content, b.version AS brs_version,
               b.status AS brs_status
        FROM automation_test_tasks t
        JOIN change_requests cr ON cr.id = t.cr_id
        LEFT JOIN projects p ON p.id = cr.project_id
        LEFT JOIN cr_brs_documents b ON b.id = t.brs_id
        WHERE t.id = $1
        """,
        task_id,
    )
    if not row:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Task test không tồn tại"})
    return row


def _parse_cases_json(raw: str) -> list[dict]:
    """Parse mảng JSON test case do AI trả về (chấp nhận có markdown fence)."""
    text = raw.strip()
    fence = re.search(r"```(?:json)?\s*(.*?)```", text, re.DOTALL)
    if fence:
        text = fence.group(1).strip()
    start, end = text.find("["), text.rfind("]")
    if start == -1 or end == -1:
        raise HTTPException(
            502,
            detail={
                "code": "AI_FORMAT_ERROR",
                "message": "AI không trả về mảng JSON test case. Kiểm tra lại skill gen_test_case.",
                "raw_preview": raw[:500],
            },
        )
    try:
        parsed = json.loads(text[start:end + 1])
    except json.JSONDecodeError as exc:
        raise HTTPException(
            502,
            detail={
                "code": "AI_FORMAT_ERROR",
                "message": f"JSON test case không hợp lệ: {exc}",
                "raw_preview": raw[:500],
            },
        )
    if not isinstance(parsed, list) or not parsed:
        raise HTTPException(
            502,
            detail={"code": "AI_FORMAT_ERROR", "message": "AI trả về danh sách test case rỗng."},
        )
    return parsed


# ── Task ─────────────────────────────────────────────────────────────────────
@router.get("/tasks")
async def list_tasks(
    user: CurrentUser,
    status: Optional[str] = None,
    project_id: Optional[str] = None,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    conditions: list[str] = []
    params: list = []
    if status:
        params.append(status)
        conditions.append(f"t.status = ${len(params)}")
    if project_id:
        params.append(project_id)
        conditions.append(f"cr.project_id = ${len(params)}")
    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    rows = await db.fetch(
        f"""
        SELECT t.*, cr.request_code, cr.title AS cr_title, cr.priority AS cr_priority,
               p.code AS project_code, p.name AS project_name,
               b.version AS brs_version, b.status AS brs_status,
               (SELECT COUNT(*) FROM automation_test_cases c WHERE c.task_id = t.id) AS case_count,
               (SELECT COUNT(*) FROM automation_test_cases c
                 WHERE c.task_id = t.id AND c.studio_tc_id IS NOT NULL) AS mapped_count,
               (SELECT COUNT(*) FROM automation_test_runs r WHERE r.task_id = t.id) AS run_count
        FROM automation_test_tasks t
        JOIN change_requests cr ON cr.id = t.cr_id
        LEFT JOIN projects p ON p.id = cr.project_id
        LEFT JOIN cr_brs_documents b ON b.id = t.brs_id
        {where}
        ORDER BY t.updated_at DESC
        """,
        *params,
    )
    return {"data": [dict(r) for r in rows]}


@router.get("/tasks/{task_id}")
async def get_task(
    user: CurrentUser,
    task_id: str,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    task = await _get_task_or_404(db, task_id)
    cases = await db.fetch(
        "SELECT * FROM automation_test_cases WHERE task_id = $1 ORDER BY sort_order, code",
        task_id,
    )
    runs = await db.fetch(
        """
        SELECT id, run_ref, summary, created_by, created_at,
               LENGTH(report_content) AS report_length
        FROM automation_test_runs WHERE task_id = $1 ORDER BY created_at DESC
        """,
        task_id,
    )
    data = dict(task)
    data.pop("brs_content", None)  # nội dung BRS dài — lấy riêng khi cần
    return {
        "data": data,
        "cases": [dict(c) for c in cases],
        "runs": [dict(r) for r in runs],
    }


@router.post("/tasks/{task_id}/generate-cases", status_code=201)
async def generate_cases(
    user: CurrentUser,
    task_id: str,
    body: GenerateCasesRequest,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    """[Gen testcase] — AI sinh test case từ BRS đã duyệt của CR."""
    assert_skill_for_step("gen_test_case", body.skill_code)
    task = await _get_task_or_404(db, task_id)
    if task["status"] == "closed":
        raise HTTPException(
            409, detail={"code": "TASK_CLOSED", "message": "Task test đã đóng."}
        )
    if not task["brs_content"]:
        raise HTTPException(
            409,
            detail={
                "code": "BRS_MISSING",
                "message": "Task chưa gắn BRS có nội dung. Duyệt BRS của CR trước.",
            },
        )
    if task["brs_status"] not in ("approved", "golive"):
        raise HTTPException(
            409,
            detail={
                "code": "BRS_NOT_APPROVED",
                "message": "BRS phải được duyệt trước khi sinh test case.",
            },
        )

    prompt = (
        f"=== CHANGE REQUEST ===\n{task['request_code']} — {task['cr_title']}\n"
        f"{task['cr_description'] or ''}\n\n"
        f"=== BRS ĐÃ DUYỆT (v{task['brs_version']}) ===\n{task['brs_content']}\n\n"
        f"{('Ghi chú của QA: ' + body.note) if body.note.strip() else ''}\n"
        "Sinh bộ test case cho thay đổi trên."
    )
    raw = await run_skill(db, body.skill_code, prompt)
    parsed = _parse_cases_json(raw)

    # Giữ lại case đã map studio (đã record script) — chỉ thay case chưa map
    kept = await db.fetch(
        "SELECT code FROM automation_test_cases WHERE task_id = $1 AND studio_tc_id IS NOT NULL",
        task_id,
    )
    kept_codes = {r["code"] for r in kept}

    created = 0
    async with db.transaction():
        # Sinh lại sẽ THAY test case chưa map script. Đếm và báo lại số bị thay: nếu QA đã
        # sửa tay nội dung một case chưa map thì bản sửa đó mất, phải cho họ biết.
        replaced_row = await db.fetchval(
            "SELECT COUNT(*) FROM automation_test_cases "
            "WHERE task_id = $1 AND studio_tc_id IS NULL",
            task_id,
        )
        replaced = replaced_row or 0
        await db.execute(
            "DELETE FROM automation_test_cases WHERE task_id = $1 AND studio_tc_id IS NULL",
            task_id,
        )
        for idx, item in enumerate(parsed):
            code = str(item.get("code") or f"TC-{idx + 1:02d}").strip()[:40]
            if code in kept_codes:
                continue
            priority = str(item.get("priority") or "medium").lower()
            if priority not in VALID_PRIORITIES:
                priority = "medium"
            await db.execute(
                """
                INSERT INTO automation_test_cases
                    (id, task_id, code, title, precondition, steps, expected,
                     priority, status, sort_order, created_by)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,'ready',$9,$10)
                """,
                str(uuid4()), task_id, code,
                str(item.get("title") or code)[:300],
                str(item.get("precondition") or ""),
                str(item.get("steps") or ""),
                str(item.get("expected") or ""),
                priority, idx, user.sub,
            )
            created += 1
        await db.execute(
            "UPDATE automation_test_tasks SET status = 'cases_generated', updated_at = NOW() "
            "WHERE id = $1 AND status = 'need_test'",
            task_id,
        )

    await log_audit(
        db=db, entity_type="automation_test_tasks", entity_id=task_id, action="CREATE",
        changed_by=user.sub, new_values={"generated_cases": created, "kept_mapped": len(kept_codes),
                    "replaced_unmapped": replaced},
    )
    cases = await db.fetch(
        "SELECT * FROM automation_test_cases WHERE task_id = $1 ORDER BY sort_order, code", task_id
    )
    return {
        "data": [dict(c) for c in cases],
        "meta": {
            "created": created,
            "kept": len(kept_codes),
            "replaced": replaced,
            "message": (
                f"Đã thay {replaced} test case chưa map script"
                if replaced else "Không có test case nào bị thay"
            ) + (f", giữ {len(kept_codes)} case đã map." if kept_codes else "."),
        },
    }


# ── Test case CRUD ───────────────────────────────────────────────────────────
@router.post("/tasks/{task_id}/cases", status_code=201)
async def create_case(
    user: CurrentUser,
    task_id: str,
    body: TestCaseCreate,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    await _get_task_or_404(db, task_id)
    if body.priority not in VALID_PRIORITIES:
        raise HTTPException(
            400, detail={"code": "VALIDATION_ERROR", "message": "Độ ưu tiên không hợp lệ"}
        )
    max_order = await db.fetchval(
        "SELECT COALESCE(MAX(sort_order), -1) FROM automation_test_cases WHERE task_id = $1",
        task_id,
    )
    row = await db.fetchrow(
        """
        INSERT INTO automation_test_cases
            (id, task_id, code, title, precondition, steps, expected, priority,
             status, sort_order, created_by)
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,'ready',$9,$10) RETURNING *
        """,
        str(uuid4()), task_id, body.code, body.title, body.precondition,
        body.steps, body.expected, body.priority, max_order + 1, user.sub,
    )
    return {"data": dict(row)}


@router.put("/cases/{case_id}")
async def update_case(
    user: CurrentUser,
    case_id: str,
    body: TestCaseUpdate,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    existing = await db.fetchrow("SELECT * FROM automation_test_cases WHERE id = $1", case_id)
    if not existing:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Test case không tồn tại"})

    updates = body.model_dump(exclude_none=True)
    if not updates:
        raise HTTPException(
            400, detail={"code": "VALIDATION_ERROR", "message": "Không có thay đổi nào"}
        )
    if "priority" in updates and updates["priority"] not in VALID_PRIORITIES:
        raise HTTPException(
            400, detail={"code": "VALIDATION_ERROR", "message": "Độ ưu tiên không hợp lệ"}
        )
    # Gán studio_tc_id = đã map script automation
    if updates.get("studio_tc_id") and "status" not in updates:
        updates["status"] = "mapped"

    set_parts = [f"{k} = ${i + 2}" for i, k in enumerate(updates.keys())]
    row = await db.fetchrow(
        f"UPDATE automation_test_cases SET {', '.join(set_parts)}, updated_at = NOW() "
        f"WHERE id = $1 RETURNING *",
        case_id, *updates.values(),
    )
    return {"data": dict(row)}


@router.delete("/cases/{case_id}", status_code=204, response_class=Response)
async def delete_case(
    user: CurrentUser,
    case_id: str,
    db: asyncpg.Connection = Depends(get_db),
) -> Response:
    result = await db.execute("DELETE FROM automation_test_cases WHERE id = $1", case_id)
    if result.endswith("0"):
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Test case không tồn tại"})
    return Response(status_code=204)


# ── Kết quả chạy ─────────────────────────────────────────────────────────────
@router.post("/tasks/{task_id}/runs", status_code=201)
async def import_run(
    user: CurrentUser,
    task_id: str,
    body: RunImport,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    """Import kết quả một lượt chạy từ Capture Studio và cập nhật trạng thái test case."""
    task = await _get_task_or_404(db, task_id)
    if task["status"] == "closed":
        raise HTTPException(409, detail={"code": "TASK_CLOSED", "message": "Task test đã đóng."})

    run_id = str(uuid4())
    results = body.summary.get("cases") or []
    updated = 0
    async with db.transaction():
        await db.execute(
            """
            INSERT INTO automation_test_runs (id, task_id, run_ref, summary, created_by)
            VALUES ($1, $2, $3, $4, $5)
            """,
            run_id, task_id, body.run_ref, body.summary, user.sub,
        )
        for item in results:
            studio_id = item.get("studio_tc_id") or item.get("testcaseId") or item.get("id")
            outcome = str(item.get("status") or "").lower()
            if not studio_id or outcome not in ("passed", "failed"):
                continue
            res = await db.execute(
                "UPDATE automation_test_cases SET status = $3, updated_at = NOW() "
                "WHERE task_id = $1 AND studio_tc_id = $2",
                task_id, str(studio_id), outcome,
            )
            if not res.endswith("0"):
                updated += 1
        await db.execute(
            "UPDATE automation_test_tasks SET status = 'in_progress', updated_at = NOW() "
            "WHERE id = $1 AND status <> 'closed'",
            task_id,
        )
    await log_audit(
        db=db, entity_type="automation_test_runs", entity_id=run_id, action="CREATE",
        changed_by=user.sub, new_values={"task_id": task_id, "cases_updated": updated},
    )
    row = await db.fetchrow("SELECT * FROM automation_test_runs WHERE id = $1", run_id)
    return {"data": dict(row), "meta": {"cases_updated": updated}}


@router.post("/runs/{run_id}/generate-report")
async def generate_report(
    user: CurrentUser,
    run_id: str,
    body: ReportRequest,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    """[Gen report] — AI tổng hợp báo cáo kết quả từ một lượt chạy."""
    assert_skill_for_step("gen_test_report", body.skill_code)
    run = await db.fetchrow("SELECT * FROM automation_test_runs WHERE id = $1", run_id)
    if not run:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Lượt chạy không tồn tại"})
    task = await _get_task_or_404(db, str(run["task_id"]))
    cases = await db.fetch(
        "SELECT code, title, priority, status, expected, studio_tc_id FROM automation_test_cases "
        "WHERE task_id = $1 ORDER BY sort_order, code",
        run["task_id"],
    )

    summary = run["summary"] if isinstance(run["summary"], dict) else json.loads(run["summary"] or "{}")

    # Kết quả PHẢI lấy từ chính lượt chạy này, không lấy cột status của test case.
    #
    # automation_test_cases.status là kết quả của LƯỢT CHẠY GẦN NHẤT. Nếu chạy lại rồi mới
    # sinh báo cáo cho lượt cũ, báo cáo sẽ mang số liệu của lượt mới — sai lệch mà không có
    # dấu hiệu nào. Ngoài ra skill yêu cầu nêu "Ghi nhận" của case fail, dữ liệu đó chỉ có
    # trong summary do Capture Studio đẩy sang; không truyền vào thì AI buộc phải bịa.
    run_results: dict[str, dict] = {}
    for item in (summary.get("cases") or []):
        key = str(item.get("studio_tc_id") or item.get("testcaseId") or item.get("id") or "")
        if key:
            run_results[key] = item

    case_lines_parts: list[str] = []
    missing_in_run = 0
    for c in cases:
        res = run_results.get(str(c["studio_tc_id"] or ""))
        if res:
            outcome = str(res.get("status") or "không rõ")
            note = (res.get("error") or res.get("message") or res.get("note") or "").strip()
            duration = res.get("duration_ms")
        else:
            missing_in_run += 1
            outcome = "KHÔNG CÓ TRONG LƯỢT CHẠY NÀY"
            note = ""
            duration = None
        line = f"- {c['code']} | {c['title']} | ưu tiên {c['priority']} | kết quả: {outcome}"
        if duration:
            line += f" | {duration} ms"
        line += f" | ghi nhận: {note if note else 'không có ghi nhận lỗi từ công cụ chạy'}"
        case_lines_parts.append(line)
    case_lines = "\n".join(case_lines_parts)
    if missing_in_run:
        case_lines += (
            f"\n\nLƯU Ý: {missing_in_run} test case không có kết quả trong lượt chạy này "
            "(chưa map script hoặc chưa được chọn để chạy). Phải nêu rõ trong báo cáo là "
            "CHƯA CHẠY, tuyệt đối không suy ra là đạt."
        )
    prompt = (
        f"=== CHANGE REQUEST ===\n{task['request_code']} — {task['cr_title']}\n"
        f"Dự án: {task['project_code'] or ''} {task['project_name'] or ''}\n"
        f"BRS: v{task['brs_version'] or '-'}\n\n"
        f"=== LƯỢT CHẠY ===\nMã lượt chạy: {run['run_ref'] or run_id}\n"
        f"Thời điểm: {run['created_at']}\n"
        f"Số liệu tổng hợp: {json.dumps(summary, ensure_ascii=False)}\n\n"
        f"=== DANH SÁCH TEST CASE ===\n{case_lines}\n\n"
        "Viết báo cáo kết quả kiểm thử cho lượt chạy trên."
    )
    content = await run_skill(db, body.skill_code, prompt)
    row = await db.fetchrow(
        "UPDATE automation_test_runs SET report_content = $2 WHERE id = $1 RETURNING *",
        run_id, content,
    )
    return {"data": dict(row)}


@router.get("/runs/{run_id}")
async def get_run(
    user: CurrentUser,
    run_id: str,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    row = await db.fetchrow("SELECT * FROM automation_test_runs WHERE id = $1", run_id)
    if not row:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Lượt chạy không tồn tại"})
    return {"data": dict(row)}


@router.post("/tasks/{task_id}/close")
async def close_task(
    user: CurrentUser,
    task_id: str,
    db: asyncpg.Connection = Depends(get_db),
) -> dict:
    """[Close] — đóng công việc test của CR sau khi QA thấy kết quả đạt."""
    task = await _get_task_or_404(db, task_id)
    if task["status"] == "closed":
        raise HTTPException(409, detail={"code": "TASK_CLOSED", "message": "Task test đã đóng."})
    run_count = await db.fetchval(
        "SELECT COUNT(*) FROM automation_test_runs WHERE task_id = $1", task_id
    )
    if not run_count:
        raise HTTPException(
            409,
            detail={
                "code": "NO_RUN",
                "message": "Chưa có lượt chạy nào — không đóng được task test.",
            },
        )
    row = await db.fetchrow(
        """
        UPDATE automation_test_tasks
        SET status = 'closed', closed_by = $2, closed_at = NOW(), updated_at = NOW()
        WHERE id = $1 RETURNING *
        """,
        task_id, user.sub,
    )
    await db.execute(
        """
        INSERT INTO request_history
            (ref_type, ref_id, action, actor, from_status, to_status, comment)
        VALUES ('cr', $1::uuid, 'TEST_CLOSED', $2, $3, 'closed', $4)
        """,
        str(task["cr_id"]), user.sub, task["status"],
        f"Đóng công việc automation test ({run_count} lượt chạy)",
    )
    await log_audit(
        db=db, entity_type="automation_test_tasks", entity_id=task_id, action="STATUS_CHANGE",
        changed_by=user.sub, new_values={"status": "closed"},
    )
    return {"data": dict(row)}


# ── Export ───────────────────────────────────────────────────────────────────
@router.get("/tasks/{task_id}/export")
async def export_task(
    user: CurrentUser,
    task_id: str,
    db: asyncpg.Connection = Depends(get_db),
) -> StreamingResponse:
    """Xuất test case + kết quả chạy + báo cáo ra file Excel."""
    task = await _get_task_or_404(db, task_id)
    cases = await db.fetch(
        "SELECT * FROM automation_test_cases WHERE task_id = $1 ORDER BY sort_order, code", task_id
    )
    runs = await db.fetch(
        "SELECT * FROM automation_test_runs WHERE task_id = $1 ORDER BY created_at DESC", task_id
    )

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="155EEF")
    wrap = Alignment(vertical="top", wrap_text=True)

    # Sheet 1 — Test case
    ws = wb.active
    ws.title = "Test case"
    ws.append([f"Automation Test — {task['request_code']}: {task['cr_title']}"])
    ws.append([f"Dự án: {task['project_code'] or ''} {task['project_name'] or ''}"
               f"   |   BRS: v{task['brs_version'] or '-'} ({task['brs_status'] or '-'})"
               f"   |   Trạng thái: {task['status']}"])
    ws.append([])
    headers = ["Mã", "Tiêu đề", "Điều kiện tiên quyết", "Các bước", "Kết quả mong đợi",
               "Ưu tiên", "Kết quả", "Studio TC"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
    for c in cases:
        ws.append([c["code"], c["title"], c["precondition"], c["steps"], c["expected"],
                   c["priority"], c["status"], c["studio_tc_id"] or ""])
    for col, width in zip("ABCDEFGH", [12, 40, 30, 50, 40, 12, 12, 24]):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = wrap

    # Sheet 2 — Lượt chạy
    ws2 = wb.create_sheet("Lượt chạy")
    run_headers = ["Thời điểm", "Mã lượt chạy", "Tổng", "Đạt", "Không đạt", "Người chạy"]
    ws2.append(run_headers)
    for col in range(1, len(run_headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    for r in runs:
        summary = r["summary"] if isinstance(r["summary"], dict) else json.loads(r["summary"] or "{}")
        ws2.append([
            r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "",
            r["run_ref"] or "",
            summary.get("total", ""), summary.get("passed", ""), summary.get("failed", ""),
            r["created_by"] or "",
        ])
    for col, width in zip("ABCDEF", [20, 28, 10, 10, 12, 18]):
        ws2.column_dimensions[col].width = width

    # Sheet 3 — Báo cáo mới nhất
    ws3 = wb.create_sheet("Báo cáo")
    latest = next((r for r in runs if (r["report_content"] or "").strip()), None)
    if latest:
        ws3.append([f"Báo cáo lượt chạy {latest['run_ref'] or latest['id']}"])
        ws3.cell(row=1, column=1).font = Font(bold=True, size=13)
        ws3.append([])
        for line in (latest["report_content"] or "").splitlines():
            ws3.append([line])
    else:
        ws3.append(["Chưa có báo cáo"])
        ws3.append(["Sinh báo cáo bằng nút [Gen report] sau khi có lượt chạy."])
    ws3.column_dimensions["A"].width = 120

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"automation_test_{task['request_code']}_{stamp}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Sinh test case / báo cáo có tường thuật tiến độ (SSE) ────────────────────

GEN_CASES_STEPS = [
    {"id": "task",  "label": "Kiểm tra công việc test và BRS"},
    {"id": "keep",  "label": "Rà test case đã map script"},
    {"id": "ai",    "label": "Claude sinh bộ test case"},
    {"id": "save",  "label": "Lưu test case"},
]

GEN_REPORT_STEPS = [
    {"id": "run",     "label": "Đọc lượt chạy"},
    {"id": "results", "label": "Ghép kết quả của CHÍNH lượt chạy này"},
    {"id": "ai",      "label": "Claude viết báo cáo"},
    {"id": "save",    "label": "Lưu báo cáo vào lượt chạy"},
]


def _at_hook(emit):
    async def h(ev: dict) -> None:
        emit(ev)
    return h


@router.post("/tasks/{task_id}/generate-cases/stream")
async def generate_cases_stream(
    user: CurrentUser,
    task_id: str,
    body: GenerateCasesRequest,
):
    """[Gen testcase] có tường thuật — cùng kết quả với POST /tasks/{id}/generate-cases."""
    actor = user.sub

    async def work(db: asyncpg.Connection, emit) -> dict:
        assert_skill_for_step("gen_test_case", body.skill_code)
        task = await _get_task_or_404(db, task_id)
        if task["status"] == "closed":
            raise HTTPException(409, detail={"code": "TASK_CLOSED",
                                             "message": "Task test đã đóng."})
        if not task["brs_content"]:
            raise HTTPException(409, detail={
                "code": "BRS_MISSING",
                "message": "Task chưa gắn BRS có nội dung. Duyệt BRS của CR trước.",
            })
        if task["brs_status"] not in ("approved", "golive"):
            raise HTTPException(409, detail={
                "code": "BRS_NOT_APPROVED",
                "message": "BRS phải được duyệt trước khi sinh test case.",
            })
        step_done(emit, "task", f"{task['request_code']} · BRS v{task['brs_version']} "
                                f"({task['brs_status']}) · "
                                f"{vi_num(len(task['brs_content'] or ''))} ký tự")

        step_running(emit, "keep")
        kept = await db.fetch(
            "SELECT code FROM automation_test_cases "
            "WHERE task_id = $1 AND studio_tc_id IS NOT NULL",
            task_id,
        )
        kept_codes = {r["code"] for r in kept}
        will_replace = await db.fetchval(
            "SELECT COUNT(*) FROM automation_test_cases "
            "WHERE task_id = $1 AND studio_tc_id IS NULL",
            task_id,
        ) or 0
        step_done(emit, "keep",
                  f"giữ {len(kept_codes)} case đã map script · "
                  f"sẽ thay {will_replace} case chưa map"
                  + (" (bản QA sửa tay ở các case này sẽ mất)" if will_replace else ""))

        prompt = (
            f"=== CHANGE REQUEST ===\n{task['request_code']} — {task['cr_title']}\n"
            f"{task['cr_description'] or ''}\n\n"
            f"=== BRS ĐÃ DUYỆT (v{task['brs_version']}) ===\n{task['brs_content']}\n\n"
            f"{('Ghi chú của QA: ' + body.note) if body.note.strip() else ''}\n"
            "Sinh bộ test case cho thay đổi trên."
        )
        step_running(emit, "ai")
        raw = await run_skill(db, body.skill_code, prompt, on_event=_at_hook(emit))
        parsed = _parse_cases_json(raw)
        step_done(emit, "ai", f"{len(parsed)} test case trong phản hồi")

        step_running(emit, "save")
        created = 0
        async with db.transaction():
            replaced = will_replace
            await db.execute(
                "DELETE FROM automation_test_cases WHERE task_id = $1 AND studio_tc_id IS NULL",
                task_id,
            )
            for idx, item in enumerate(parsed):
                code = str(item.get("code") or f"TC-{idx + 1:02d}").strip()[:40]
                if code in kept_codes:
                    continue
                priority = str(item.get("priority") or "medium").lower()
                if priority not in VALID_PRIORITIES:
                    priority = "medium"
                await db.execute(
                    """
                    INSERT INTO automation_test_cases
                        (id, task_id, code, title, precondition, steps, expected,
                         priority, status, sort_order, created_by)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,'ready',$9,$10)
                    """,
                    str(uuid4()), task_id, code,
                    str(item.get("title") or code)[:300],
                    str(item.get("precondition") or ""),
                    str(item.get("steps") or ""),
                    str(item.get("expected") or ""),
                    priority, idx, actor,
                )
                created += 1
            await db.execute(
                "UPDATE automation_test_tasks SET status = 'cases_generated', updated_at = NOW() "
                "WHERE id = $1 AND status = 'need_test'",
                task_id,
            )
        await log_audit(
            db=db, entity_type="automation_test_tasks", entity_id=task_id, action="CREATE",
            changed_by=actor,
            new_values={"generated_cases": created, "kept_mapped": len(kept_codes),
                        "replaced_unmapped": replaced},
        )
        cases = await db.fetch(
            "SELECT * FROM automation_test_cases WHERE task_id = $1 ORDER BY sort_order, code",
            task_id,
        )
        step_done(emit, "save", f"tạo {created} · giữ {len(kept_codes)} · thay {replaced}")
        return {
            "data": [dict(c) for c in cases],
            "meta": {
                "created": created, "kept": len(kept_codes), "replaced": replaced,
                "message": (
                    f"Đã thay {replaced} test case chưa map script"
                    if replaced else "Không có test case nào bị thay"
                ) + (f", giữ {len(kept_codes)} case đã map." if kept_codes else "."),
            },
        }

    return sse_response(GEN_CASES_STEPS, work)


@router.post("/runs/{run_id}/generate-report/stream")
async def generate_report_stream(
    user: CurrentUser,
    run_id: str,
    body: ReportRequest,
):
    """[Gen report] có tường thuật — cùng kết quả với POST /runs/{id}/generate-report."""

    async def work(db: asyncpg.Connection, emit) -> dict:
        assert_skill_for_step("gen_test_report", body.skill_code)
        run = await db.fetchrow("SELECT * FROM automation_test_runs WHERE id = $1", run_id)
        if not run:
            raise HTTPException(404, detail={"code": "NOT_FOUND",
                                             "message": "Lượt chạy không tồn tại"})
        task = await _get_task_or_404(db, str(run["task_id"]))
        step_done(emit, "run", f"{run['run_ref'] or run_id} · {task['request_code']}")

        step_running(emit, "results")
        cases = await db.fetch(
            "SELECT code, title, priority, status, expected, studio_tc_id "
            "FROM automation_test_cases WHERE task_id = $1 ORDER BY sort_order, code",
            run["task_id"],
        )
        summary = run["summary"] if isinstance(run["summary"], dict) \
            else json.loads(run["summary"] or "{}")
        run_results: dict[str, dict] = {}
        for item in (summary.get("cases") or []):
            key = str(item.get("studio_tc_id") or item.get("testcaseId") or item.get("id") or "")
            if key:
                run_results[key] = item

        case_lines_parts: list[str] = []
        missing_in_run = 0
        for c in cases:
            res = run_results.get(str(c["studio_tc_id"] or ""))
            if res:
                outcome = str(res.get("status") or "không rõ")
                note = (res.get("error") or res.get("message") or res.get("note") or "").strip()
                duration = res.get("duration_ms")
            else:
                missing_in_run += 1
                outcome = "KHÔNG CÓ TRONG LƯỢT CHẠY NÀY"
                note = ""
                duration = None
            line = (f"- {c['code']} | {c['title']} | ưu tiên {c['priority']} | "
                    f"kết quả: {outcome}")
            if duration:
                line += f" | {duration} ms"
            line += f" | ghi nhận: {note if note else 'không có ghi nhận lỗi từ công cụ chạy'}"
            case_lines_parts.append(line)
        case_lines = "\n".join(case_lines_parts)
        if missing_in_run:
            case_lines += (
                f"\n\nLƯU Ý: {missing_in_run} test case không có kết quả trong lượt chạy này "
                "(chưa map script hoặc chưa được chọn để chạy). Phải nêu rõ trong báo cáo là "
                "CHƯA CHẠY, tuyệt đối không suy ra là đạt."
            )
        step_done(emit, "results",
                  f"{len(cases)} case · {len(cases) - missing_in_run} có kết quả trong lượt này"
                  + (f" · {missing_in_run} CHƯA CHẠY" if missing_in_run else ""))

        prompt = (
            f"=== CHANGE REQUEST ===\n{task['request_code']} — {task['cr_title']}\n"
            f"Dự án: {task['project_code'] or ''} {task['project_name'] or ''}\n"
            f"BRS: v{task['brs_version'] or '-'}\n\n"
            f"=== LƯỢT CHẠY ===\nMã lượt chạy: {run['run_ref'] or run_id}\n"
            f"Thời điểm: {run['created_at']}\n"
            f"Số liệu tổng hợp: {json.dumps(summary, ensure_ascii=False)}\n\n"
            f"=== DANH SÁCH TEST CASE ===\n{case_lines}\n\n"
            "Viết báo cáo kết quả kiểm thử cho lượt chạy trên."
        )
        step_running(emit, "ai")
        content = await run_skill(db, body.skill_code, prompt, on_event=_at_hook(emit))
        step_done(emit, "ai")

        step_running(emit, "save")
        row = await db.fetchrow(
            "UPDATE automation_test_runs SET report_content = $2 WHERE id = $1 RETURNING *",
            run_id, content,
        )
        step_done(emit, "save", f"{vi_num(len(content))} ký tự")
        return {"data": dict(row)}

    return sse_response(GEN_REPORT_STEPS, work)
