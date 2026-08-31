"""
GNM Export Service — parses Markdown docs (BRD/FRS/API spec/ETL) into
GNM 3-layer hierarchy and generates an Excel file from the GNM.xlsx template.

GNM layers:
  Layer 1 — Root Title    : H1 or filename
  Layer 2 — Management    : H2 sections  (4–8 items)
  Layer 3 — Sub-items     : H3 or bullets under each H2  (3–7 items each)
  Col H   — Detail        : first substantial paragraph (60–220 chars)
"""
from __future__ import annotations

import io
import math
import os
import re
import shutil
import tempfile
from copy import copy
from datetime import date
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font

# ── Template path ──────────────────────────────────────────────────────────────
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_TEMPLATE = os.path.normpath(
    os.path.join(_THIS_DIR, "..", "..", "..", "..", "docs", "brd", "export-GNM", "GNM.xlsx")
)
GNM_TEMPLATE_PATH: str = os.getenv("GNM_TEMPLATE_PATH", _DEFAULT_TEMPLATE)

# ── Style constants ───────────────────────────────────────────────────────────
_FONT_NAME  = "Myriad Pro"
_DATA_FONT  = Font(name=_FONT_NAME, size=11)
_TITLE_FONT = Font(name=_FONT_NAME, size=14, bold=True)
_EMPTY_BDR  = Border()


# ── Markdown parser ───────────────────────────────────────────────────────────

def _clean_heading(raw: str) -> str:
    """Strip leading numbering (1., 1.1, FR-001 etc.) and markdown symbols."""
    h = raw.strip().lstrip("#").strip()
    h = re.sub(r"^[\d\.]+\s+", "", h)          # "1.2 Title" → "Title"
    h = re.sub(r"^\*+|\*+$", "", h).strip()    # bold markers
    return h[:80]


def _extract_detail(lines: List[str], min_len: int = 60, max_len: int = 220) -> str:
    """Build a detail string from content lines (≥60 chars, ≤220 chars)."""
    # Prefer real paragraphs; skip blank lines and code fences
    paragraphs: List[str] = []
    buf: List[str] = []
    in_code = False

    for ln in lines:
        if ln.strip().startswith("```"):
            in_code = not in_code
            continue
        if in_code:
            continue
        stripped = ln.strip()
        if stripped:
            buf.append(stripped)
        else:
            if buf:
                paragraphs.append(" ".join(buf))
                buf = []
    if buf:
        paragraphs.append(" ".join(buf))

    # Join paragraphs until we have enough text
    text = ""
    for p in paragraphs:
        # Strip markdown syntax chars
        clean = re.sub(r"[#*`>\-_|]", "", p).strip()
        # Remove bullet markers
        clean = re.sub(r"^\s*[\-\*\+]\s+", "", clean)
        if len(clean) < 5:
            continue
        text = (text + " " + clean).strip() if text else clean
        if len(text) >= min_len:
            break

    if not text:
        text = "Xem tài liệu gốc để biết thêm chi tiết về mục này."

    # Pad if still too short
    if len(text) < min_len:
        text = text + " — Tham khảo tài liệu đính kèm để biết thêm thông tin."

    return text[:max_len]


def _bullets_to_subs(lines: List[str]) -> List[Tuple[str, str]]:
    """Extract bullet list items from lines as (name, detail) tuples."""
    subs: List[Tuple[str, str]] = []
    current_item: str | None = None
    current_lines: List[str] = []

    for ln in lines:
        m = re.match(r"^\s*[-*+]\s+(.+)$", ln)
        m_num = re.match(r"^\s*\d+\.\s+(.+)$", ln)
        matched = m or m_num
        if matched:
            if current_item:
                subs.append((current_item, _extract_detail(current_lines or [current_item])))
            current_item = _clean_heading(matched.group(1))
            current_lines = []
        elif current_item and ln.strip():
            current_lines.append(ln)

    if current_item:
        subs.append((current_item, _extract_detail(current_lines or [current_item])))

    return subs


def parse_markdown_to_gnm(content: str, filename: str = "Document") -> Dict:
    """
    Parse markdown content into GNM 3-layer dict:
      {
        "root_title": str,
        "management": [str, ...],           # Layer 2
        "sub_items":  {str: [(str,str)]},   # Layer 3: {mgmt: [(name, detail)]}
      }
    """
    lines = content.split("\n")

    # ── Layer 1: Root Title ──────────────────────────────────────────────────
    root_title = ""
    for ln in lines:
        if re.match(r"^# [^#]", ln):
            root_title = _clean_heading(ln[2:])
            break
    if not root_title:
        root_title = os.path.splitext(filename)[0].replace("_", " ").replace("-", " ").strip()

    # ── Layer 2 / 3: H2 → Management, H3 → Sub-items ────────────────────────
    management: List[str] = []
    sub_items: Dict[str, List[Tuple[str, str]]] = {}

    current_h2: str | None = None
    current_h3: str | None = None
    h3_lines: List[str] = []
    h2_lines: List[str] = []    # lines between H2 and first H3

    def _flush_h3():
        nonlocal current_h3, h3_lines
        if current_h3 and current_h2:
            detail = _extract_detail(h3_lines or [current_h3])
            sub_items.setdefault(current_h2, []).append((current_h3, detail))
        current_h3 = None
        h3_lines = []

    def _flush_h2_bullets():
        """If no H3s found under current_h2, extract bullets from h2_lines."""
        if current_h2 and not sub_items.get(current_h2):
            subs = _bullets_to_subs(h2_lines)
            if subs:
                sub_items[current_h2] = subs
        h2_lines.clear()

    for ln in lines:
        if re.match(r"^## [^#]", ln):
            _flush_h3()
            _flush_h2_bullets()
            current_h2 = _clean_heading(ln[3:])
            if current_h2 and current_h2 not in management:
                management.append(current_h2)
                sub_items.setdefault(current_h2, [])
            current_h3 = None
            h3_lines = []
            h2_lines = []
        elif re.match(r"^### [^#]", ln):
            _flush_h3()
            h2_lines = []        # stop collecting H2-level bullets once H3 exists
            current_h3 = _clean_heading(ln[4:])
            h3_lines = []
        else:
            if current_h3:
                h3_lines.append(ln)
            elif current_h2:
                h2_lines.append(ln)

    _flush_h3()
    _flush_h2_bullets()

    # ── Normalise Layer 2 count (4–8) ────────────────────────────────────────
    if len(management) > 8:
        management = management[:8]

    # Ensure minimum 3 management items (create synthetic ones if needed)
    if len(management) < 3 and management:
        while len(management) < 3:
            synthetic = f"Section {len(management)+1}"
            management.append(synthetic)
            sub_items[synthetic] = [(f"{synthetic} — thông tin bổ sung",
                                     "Xem tài liệu gốc để biết thêm chi tiết về phần này.")]

    # ── Normalise Layer 3 (3–7 per item) ─────────────────────────────────────
    for mgmt in management:
        subs = sub_items.get(mgmt, [])
        if not subs:
            # Fallback: create a single sub-item from the management name
            subs = [(mgmt + " — tổng quan",
                     f"Xem tài liệu gốc để biết thêm chi tiết về phần '{mgmt}'.")]
        if len(subs) > 7:
            subs = subs[:7]
        sub_items[mgmt] = subs

    return {
        "root_title": root_title[:60],
        "management": management,
        "sub_items": sub_items,
    }


# ── Excel generator ───────────────────────────────────────────────────────────

def _copy_from_ref(src, sr: int, dst, dr: int, mc: int = 14) -> None:
    for col in range(1, mc + 1):
        s = src.cell(sr, col)
        d = dst.cell(dr, col)
        d.value = s.value
        if s.has_style:
            d.font       = copy(s.font)
            d.fill       = copy(s.fill)
            d.border     = copy(s.border)
            d.alignment  = copy(s.alignment)
            d.number_format = s.number_format


def _wipe_row(ws, r: int, mc: int = 14) -> None:
    for col in range(1, mc + 1):
        c = ws.cell(r, col)
        c.value  = None
        c.border = _EMPTY_BDR


def _set_h(ws, r: int, h: float) -> None:
    ws.row_dimensions[r].height = h


def _row_h(txt: str, base: int = 18) -> float:
    return max(base, math.ceil(len(str(txt or "")) / 100) * base + 4)


def build_gnm_excel(gnm_data: Dict, template_path: str = GNM_TEMPLATE_PATH) -> bytes:
    """
    Generate a GNM Excel file from parsed data and the GNM.xlsx template.
    Returns raw bytes of the .xlsx file.
    """
    root_title = gnm_data["root_title"]
    management  = gnm_data["management"]
    sub_items   = gnm_data["sub_items"]

    # Work in a temp dir so we don't mutate the original template
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "output.xlsx")
        shutil.copy(template_path, output_path)

        # Reference template (read-only) for style copying
        tmpl_ref = openpyxl.load_workbook(template_path, data_only=True)
        tmpl0    = tmpl_ref["0"]
        tmpl1    = tmpl_ref["1"]

        wb = openpyxl.load_workbook(output_path)

        # ── Sheet 0 ───────────────────────────────────────────────────────────
        ws0 = wb["0"]
        ws0.sheet_view.showGridLines = False

        ws0["B2"].value     = root_title
        ws0["B2"].font      = _TITLE_FONT
        ws0["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=0)

        # Header row 4: integers (avoids formula issues)
        for col, val in {2:1, 3:2, 5:1, 6:2, 7:3, 8:4, 10:1, 11:2}.items():
            ws0.cell(4, col).value = val

        n = len(management)
        for i, mgmt in enumerate(management):
            r = 8 + i
            if r > 9:
                _copy_from_ref(tmpl0, 9, ws0, r, mc=13)
            ws0.cell(r, 3).value = i + 1 if i == 0 else f"=C{r-1}+1"
            ws0.cell(r, 3).font  = _DATA_FONT
            ws0.cell(r, 5).value = "=B5" if i == 0 else None
            ws0.cell(r, 6).value = "-"
            ws0.cell(r, 7).value = mgmt
            ws0.cell(r, 7).font  = _DATA_FONT
            ws0.cell(r, 8).value = i + 1
            ws0.cell(r, 8).font  = _DATA_FONT
            _set_h(ws0, r, _row_h(mgmt))

        # Footer rows (All / spacer / Common / spacer / close)
        all_r = 8 + n
        _copy_from_ref(tmpl0, 11, ws0, all_r, mc=13)
        ws0.cell(all_r, 3).value = "All"
        ws0.cell(all_r, 3).font  = _DATA_FONT
        _set_h(ws0, all_r, 18)

        _copy_from_ref(tmpl0, 12, ws0, all_r + 1, mc=13)

        com_r = all_r + 2
        _copy_from_ref(tmpl0, 13, ws0, com_r, mc=13)
        ws0.cell(com_r, 2).value = "Common"
        ws0.cell(com_r, 2).font  = _DATA_FONT
        ws0.cell(com_r, 3).value = "-"
        ws0.cell(com_r, 3).font  = _DATA_FONT
        _set_h(ws0, com_r, 18)

        _copy_from_ref(tmpl0, 14, ws0, com_r + 1, mc=13)
        _copy_from_ref(tmpl0, 15, ws0, com_r + 2, mc=13)

        for r in range(com_r + 3, 30):
            _wipe_row(ws0, r, mc=13)

        # ── Detail sheets 1..N ────────────────────────────────────────────────
        for idx, mgmt in enumerate(management):
            sn = str(idx + 1)
            if sn in wb.sheetnames:
                ws = wb[sn]
            else:
                ws = wb.copy_worksheet(wb["1"])
                ws.title = sn
            ws.sheet_view.showGridLines = False

            ws["B2"].value     = mgmt
            ws["B2"].font      = _TITLE_FONT
            ws["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=0)

            # Header row 4
            for col, val in {2:1, 3:2, 5:1, 6:2, 7:3, 8:4, 9:5, 11:1, 12:2}.items():
                ws.cell(4, col).value = val

            subs   = sub_items.get(mgmt, [])
            n_subs = len(subs)

            for i, (name, detail) in enumerate(subs):
                r = 8 + i
                if r > 9:
                    _copy_from_ref(tmpl1, 9, ws, r)
                ws.cell(r, 3).value = i + 1 if i == 0 else f"=C{r-1}+1"
                ws.cell(r, 3).font  = _DATA_FONT
                ws.cell(r, 5).value = "=E5" if i == 0 else None
                ws.cell(r, 7).value = name
                ws.cell(r, 7).font  = _DATA_FONT
                ws.cell(r, 8).value = detail
                ws.cell(r, 8).font  = _DATA_FONT
                ws.cell(r, 8).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True, indent=1
                )
                _set_h(ws, r, _row_h(detail))

            # Footer
            all_r2 = 8 + n_subs
            _copy_from_ref(tmpl1, 28, ws, all_r2)
            ws.cell(all_r2, 3).value = "All"
            ws.cell(all_r2, 3).font  = _DATA_FONT
            _set_h(ws, all_r2, 18)

            _copy_from_ref(tmpl1, 29, ws, all_r2 + 1)

            com_r2 = all_r2 + 2
            _copy_from_ref(tmpl1, 30, ws, com_r2)
            ws.cell(com_r2, 2).value = "Common"
            ws.cell(com_r2, 2).font  = _DATA_FONT
            ws.cell(com_r2, 3).value = "-"
            ws.cell(com_r2, 3).font  = _DATA_FONT
            _set_h(ws, com_r2, 18)

            close_r = com_r2 + 1
            _copy_from_ref(tmpl1, 31, ws, close_r)
            _set_h(ws, close_r, 18)

            for r2 in range(close_r + 1, 40):
                _wipe_row(ws, r2)

        # Remove extra numbered sheets beyond management count
        for extra_sn in [str(i) for i in range(n + 1, 10)]:
            if extra_sn in wb.sheetnames:
                del wb[extra_sn]

        # Order sheets: 0, 1, 2, ...
        all_names = ["0"] + [str(i + 1) for i in range(n)]
        for target_idx, nm in enumerate(all_names):
            if nm in wb.sheetnames:
                current_idx = wb.sheetnames.index(nm)
                wb.move_sheet(nm, offset=target_idx - current_idx)

        wb.save(output_path)

        with open(output_path, "rb") as fh:
            return fh.read()
