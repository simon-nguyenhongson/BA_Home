"""
GNM Export v2 — BRD BA_Home
Sheets: 0 (overview) | 1..6 (functional areas) | 1D/2D/3D (data objects per module)

Sheet order: 0 → 1 → 1D → 2 → 2D → 3 → 3D → 4 → 5 → 6
"""
import openpyxl, math, shutil
from openpyxl.styles import Font, Alignment, Border
from copy import copy
from datetime import date
import os

# ── Content ────────────────────────────────────────────────────────────────────

ROOT_TITLE = "BA_Home — Nen Tang Quan Tri Du An IT"

# Management items và sheet names tương ứng
# SHEET_LABEL dùng hiển thị số sheet ở cột H của sheet 0
MANAGEMENT_SHEETS = [
    ("PPG System — Quan Tri Du An",          "1",  1),
    ("PPG — Data Objects",                   "1D", "1D"),
    ("BA Workflow — Quan Ly Tai Lieu",        "2",  2),
    ("BA — Data Objects",                    "2D", "2D"),
    ("Test Platform — Tu Dong Hoa Kiem Thu", "3",  3),
    ("Test — Data Objects",                  "3D", "3D"),
    ("Kien Truc & Ha Tang",                  "4",  4),
    ("Luong Nghiep Vu Chinh",                "5",  5),
    ("Van Hanh & Trien Khai",                "6",  6),
]

SUB_ITEMS = {
    # ── Sheet 1: PPG Functional ────────────────────────────────────────────────
    "PPG System — Quan Tri Du An": [
        ("Quan ly du an & trang thai",
         "Tao, cap nhat, archive du an IT voi 4 trang thai active/on_hold/completed/archived; lien ket du an voi Annual Plan; dashboard KPI tong hop so tai lieu, test coverage, pass/fail rate."),
        ("Auto-generate 9 Milestones chuan",
         "Khi tao du an, he thong tu dong sinh 9 milestone tu Kickoff den Closure, phan bo ty le theo timeline; moi milestone tu dong gan BA tasks va Test tasks tuong ung."),
        ("Quan ly thanh vien & Alias system",
         "Quan ly vai tro PM/BA/Dev/QA/PO/Stakeholder; he thong @alias tu dong map ten day du thanh vien khi parse meeting notes; moi thanh vien co alias rieng biet."),
        ("Quan ly file & File Versioning",
         "Ho tro 3 loai file: template, uploaded, external_url; luu lich su phien ban day du moi lan upload moi; tai xuong theo version cu the; copy file tu URL ngoai vao storage noi bo."),
        ("Annual Plan & Risk/Budget/KPI",
         "Ke hoach nam: state machine draft/active/closed; mo rong voi Budget (capex/opex), Resource allocation, KPI/OKR tracking, Risk Register (probability x impact = risk_score), Dependencies."),
    ],

    # ── Sheet 1D: PPG Data Objects ─────────────────────────────────────────────
    "PPG — Data Objects": [
        ("projects",
         "Bang trung tam he thong PPG: code (UNIQUE), name, status (active/on_hold/completed/archived), owner, start_date, end_date, domain_code (FK project_domains); PK la UUID; linked Annual Plan qua plan_id."),
        ("project_milestones",
         "9 milestone chuan tu dong sinh per project: name, milestone_type, start_date, end_date, status (planned/in_progress/completed/delayed), sort_order, preconditions JSONB[], done_criteria; FK projects."),
        ("project_members + project_files + file_versions",
         "Members: full_name, alias (@token), email, role, is_active. Files: file_type (template/uploaded/external_url), current_version, storage_path, external_url. Versions: history snapshot voi change_note, file_size, uploaded_by."),
        ("meeting_minutes + ppg_app_registry",
         "Meetings: raw_notes + generated_content JSONB {attendees, decisions, action_items, risks}; parse @alias. App Registry: 4 object_type (application/system/job/connection), environment JSONB, UNIQUE(project_id, code)."),
        ("ppg_annual_plans + ppg_plan_budget + ppg_plan_risks + ppg_plan_kpis",
         "Annual Plan (year, status draft/active/closed) + Budget lines (label, capex/opex, quarter, amount_planned, amount_actual, currency VND) + Risk Register (probability 1-5, impact 1-5, risk_score=P*I STORED) + KPI (target vs actual, unit)."),
        ("project_health_scores + project_stage_gates + project_stakeholders",
         "Health RAG: overall_rag/schedule_rag/budget_rag/scope_rag/team_rag/risk_rag (red/amber/green), assessed_date. Stage Gates: stage_name, gate_criteria JSONB, sign_off_by. Stakeholders: interest_level, influence_level (low/medium/high)."),
        ("project_contracts + project_licenses + project_security_info + project_operations",
         "Contracts: vendor_name, contract_type, contract_value, expiry_date, sla_details JSONB. Licenses: software_name, license_type, compliance_status. Security: data_classification (public/internal/confidential/restricted/secret), vulnerabilities JSONB. Operations: runbook, RTO/RPO, on_call_info JSONB."),
    ],

    # ── Sheet 2: BA Functional ─────────────────────────────────────────────────
    "BA Workflow — Quan Ly Tai Lieu": [
        ("Document State Machine 4 trang thai",
         "Luong phe duyet chuan: draft -> review -> approved -> archived; reject quay ve draft; luu snapshot lich su moi lan chinh sua; dam bao traceability toan bo vong doi tai lieu."),
        ("Ho tro 4 loai tai lieu nghiep vu",
         "Quan ly BRD (Business Requirements Document), BRS (Business Requirements Specification), ERD (Entity Relationship Diagram), API Specification voi versioning day du cho tung loai."),
        ("Auto-push khi Approve tai lieu",
         "Khi tai lieu duoc approve, tat ca doc types tu dong day sang PPG qua POST /sync-doc; rieng BRS day sang Test Platform qua POST /brs de trigger auto-generate test cases nen."),
        ("Stakeholder Discussions",
         "Tao thao luan gan theo tai lieu cu the voi 3 trang thai open/resolved/deferred; ghi nhan resolution notes; lam ro yeu cau nghiep vu truoc khi chuyen sang phe duyet."),
        ("BA Timeline — Gantt View",
         "Xem tien do BA tasks theo milestone dang Gantt; tracking trang thai tung tai lieu theo giai doan; ho tro BA nam bat toan bo tien do documentation trong du an."),
    ],

    # ── Sheet 2D: BA Data Objects ──────────────────────────────────────────────
    "BA — Data Objects": [
        ("requirements",
         "Bang goc luu yeu cau nghiep vu tho: project_id, title, raw_text (noi dung thô chua xu ly), status (draft), created_by; la nguon de tao documents BA chinh thuc; FK den projects (UUID)."),
        ("documents",
         "Tai lieu BA chinh thuc: req_id (FK requirements), project_id, doc_type (BRD/BRS/ERD/API), version (v1.0...), title, content JSONB (noi dung co cau truc), status (draft/review/approved/archived), reviewed_by, approved_by, pushed_at."),
        ("document_history",
         "Lich su phien ban tai lieu: doc_id (FK documents), version, changed_by, change_note, snapshot JSONB (toan bo content tai thoi diem luu), changed_at; dam bao traceability day du, khong xoa du lieu cu."),
        ("stakeholder_discussions",
         "Thao luan stakeholder: gan theo doc_id va workflow_type (ba/test); status (open/resolved/deferred); title, content, raised_by, resolution (noi dung giai quyet), resolved_by; lien ket project_id."),
        ("ba_tasks",
         "BA Tasks theo Milestone: gan milestone_id, task_type (BRD_draft/review/sign_off...), title, description, preconditions JSONB[], status (pending/in_progress/done/skipped), assigned_to, due_date, completed_at; tracking tien do BA."),
    ],

    # ── Sheet 3: Test Functional ───────────────────────────────────────────────
    "Test Platform — Tu Dong Hoa Kiem Thu": [
        ("Auto Test Case Generation tu BRS",
         "Moi business rule trong BRS tu dong sinh 1 test case + 1 Playwright script E2E; xu ly nen (background task) khong chan workflow; ket qua san sang cho Tester review ngay."),
        ("Playwright Script Auto-generation",
         "Tao Playwright E2E script tu dong cho tung business rule voi template chuan; Playwright Script Viewer tich hop trong UI de xem va copy script; san sang chay sau khi bo sung assertion."),
        ("Test Case State Machine",
         "Luong kiem thu 4 buoc: generated -> reviewed -> approved -> executed; Tester review va approve tung test case truoc khi execute; dam bao chat luong kiem soat theo quy trinh chuan."),
        ("Test Report Approval & Sync",
         "Tao bao cao kiem thu thu cong voi total/passed/failed; he thong tu tinh coverage = passed/total x 100; approve report -> tu dong push metrics ve PPG cap nhat KPI dashboard du an."),
        ("Rediff — Tai sinh Test Cases",
         "Khi BRS thay doi, trigger POST /brs/{id}/rediff de tai sinh test cases phu hop voi yeu cau moi; dam bao test coverage luon duoc dong bo voi phien ban BRS moi nhat."),
    ],

    # ── Sheet 3D: Test Data Objects ────────────────────────────────────────────
    "Test — Data Objects": [
        ("brs_sync",
         "BRS da sync tu BA: brs_id (reference toi doc_id ben BA), version, project_id (UUID), payload JSONB (toan bo BRS content: modules, business_rules[]); UNIQUE(brs_id, version); la input de auto-generate test cases."),
        ("test_cases",
         "Test Cases tu dong sinh: brs_id, brs_sync_id (FK), module (ten module nghiep vu), title, steps JSONB [{step, action, expected}], expected_result (text), playwright_script (full script), status (generated/reviewed/approved/executed)."),
        ("test_reports",
         "Bao cao kiem thu: project_id, total (tong so case), passed, failed, coverage NUMERIC(5,2) = passed/total*100, logs (text nhat ky chay test), status (generated/approved), executed_at, approved_at, pushed_at (thoi diem sync ve PPG)."),
        ("test_tasks",
         "Test Tasks theo Milestone: tuong tu ba_tasks cho Test team; milestone_id, task_type (test_plan/test_case_design/SIT/UAT/regression), preconditions JSONB[], status (pending/in_progress/done/skipped), assigned_to, due_date, completed_at."),
        ("test_results (synced to PPG)",
         "Sau khi approve test_report, metrics duoc day ve bang test_results trong PPG qua POST /sync-test: project_id, total_cases, passed, failed, coverage, executed_at; hien thi tren PPG dashboard per project."),
    ],

    # ── Sheet 4: Architecture ──────────────────────────────────────────────────
    "Kien Truc & Ha Tang": [
        ("Kien truc Microservices 3 Services",
         "PPG (:8001), BA Workflow (:8002), Test Platform (:8003) doc lap; giao tiep qua HTTP REST va sync endpoints; chia se PostgreSQL devops_hub DB; API Proxy Vite tu dong route theo prefix."),
        ("Tech Stack Backend",
         "FastAPI 0.115.5 + asyncpg 0.30.0 + Pydantic 2.10.3 lam nen tang; Python 3.12+; PostgreSQL 15+; aiokafka 0.13.0 cho event streaming; httpx 0.28.1 cho inter-service HTTP calls."),
        ("Frontend React SPA",
         "React 18.3.1 + TypeScript 5.4.5 + Vite 5.2.12 + Zustand 4.5.2;  Design System voi StatusBadge, Modal, KpiCard, ProgressBar, Toast; Vite proxy route /api/ppg, /api/ba, /api/test."),
        ("Database & File Storage",
         "PostgreSQL 15 devops_hub dung chung cho ca 3 services; asyncpg connection pool; file storage local filesystem tai uploads/{project_id}/; upgrade path ro rang sang MinIO/S3 khi can scale."),
        ("Event Streaming & Caching (Optional)",
         "Apache Kafka 3+ cho event streaming giua cac services qua aiokafka consumer; Redis 7+ cho caching; MongoDB 6+ cho document store mo rong; graceful fallback khi Kafka khong kha dung."),
    ],

    # ── Sheet 5: Business Flows ────────────────────────────────────────────────
    "Luong Nghiep Vu Chinh": [
        ("Flow 1: Tao Du An Moi",
         "POST /projects tu dong sinh 9 milestones phan bo theo timeline, tao thu muc uploads/{id}/, auto-gen BA tasks va Test tasks cho tung milestone, sinh file template markdown Charter va BRD."),
        ("Flow 2: BA Phe Duyet Tai Lieu",
         "Tao draft -> chinh sua luu history -> submit_review -> approve; khi approve: tu dong sync tat ca doc sang PPG qua /sync-doc va day BRS sang Test Platform de trigger auto-gen test cases."),
        ("Flow 3: Auto-generate Test Cases",
         "BRS duoc approve tai BA -> POST /brs toi Test Platform -> background generate_test_cases_from_brs() parse modules/rules -> moi rule sinh 1 TestCase + 1 Playwright script san sang review."),
        ("Flow 4: Bao Cao Kiem Thu",
         "Tester tao report voi total/passed/failed; he thong tinh coverage tu dong; approve report -> auto push metrics ve PPG /sync-test; dashboard du an cap nhat KPI test coverage realtime."),
        ("Flow 5: Meeting Minutes Parser",
         "POST /meetings/generate voi raw_notes chua @alias tokens; he thong map alias thanh vien, trich xuat discussion items, decisions, action items (assignee + due date), risks, suggestions dang JSON."),
    ],

    # ── Sheet 6: Operations ────────────────────────────────────────────────────
    "Van Hanh & Trien Khai": [
        ("Khoi dong & Cai dat",
         "start.bat mo 4 cua so CMD rieng cho PPG/BA/Test/Frontend; install.bat cai dependencies Python va npm; psql -f infra/init.sql khoi tao schema; health check GET /health per service."),
        ("Cau hinh moi truong (.env)",
         "DATABASE_URL, KAFKA_BOOTSTRAP, REDIS_URL cau hinh per service; BA service can PPG_SERVICE_URL=:8001 va TEST_SERVICE_URL=:8003 de auto-push; SECRET_KEY rieng moi service."),
        ("Known Issues & Gioi han (6 issues)",
         "Kafka khong ket noi local (Low); Playwright template can implement assertion (Medium); Vite CJS warning (Low); copy-from-URL thieu timeout (Low); alias parser phu thuoc member; thieu authentication (High)."),
        ("Lich su phien ban — 3 Releases",
         "v1.0.0 MVP monolith PostgreSQL; v2.0.0 microservices 3 services + auto-gen milestones + BRS sync pipeline; v3.0.0 full features: versioning, state machine, Playwright auto-gen,  Design System."),
        ("Yeu cau he thong & moi truong",
         "Bat buoc: Python 3.12+, Node.js 20+ / npm 10+, PostgreSQL 15+; tuy chon: Kafka 3+, Redis 7+, MongoDB 6+; ho tro Windows start.bat va Linux/Mac manual run voi uvicorn --reload."),
    ],
}

# ── Paths ──────────────────────────────────────────────────────────────────────
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "GNM.xlsx")
OUTPUT = os.path.normpath(os.path.join(
    os.path.dirname(__file__), "..",
    f"gnm_ba_home_v2_{date.today().strftime('%Y%m%d')}.xlsx"
))

FONT_NAME  = "Myriad Pro"
DATA_FONT  = Font(name=FONT_NAME, size=11)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True)
EMPTY_BDR  = Border()


def copy_from_ref(src, sr, dst, dr, mc=14):
    for col in range(1, mc + 1):
        s = src.cell(sr, col)
        d = dst.cell(dr, col)
        d.value = s.value
        if s.has_style:
            d.font          = copy(s.font)
            d.fill          = copy(s.fill)
            d.border        = copy(s.border)
            d.alignment     = copy(s.alignment)
            d.number_format = s.number_format


def wipe_row(ws, r, mc=14):
    for col in range(1, mc + 1):
        c = ws.cell(r, col)
        c.value  = None
        c.border = EMPTY_BDR


def set_h(ws, r, h):
    ws.row_dimensions[r].height = h


def row_h(txt, base=18):
    return max(base, math.ceil(len(str(txt or "")) / 100) * base + 4)


# ── Load template ──────────────────────────────────────────────────────────────
tmpl_ref = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
tmpl0    = tmpl_ref["0"]
tmpl1    = tmpl_ref["1"]

shutil.copy(TEMPLATE_PATH, OUTPUT)
wb = openpyxl.load_workbook(OUTPUT)

# ── Sheet 0 — Overview ─────────────────────────────────────────────────────────
ws0 = wb["0"]
ws0.sheet_view.showGridLines = False

ws0["B2"].value     = ROOT_TITLE
ws0["B2"].font      = TITLE_FONT
ws0["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=0)

# Header integers (row 4)
for col, val in {2: 1, 3: 2, 5: 1, 6: 2, 7: 3, 8: 4, 10: 1, 11: 2}.items():
    ws0.cell(4, col).value = val

n = len(MANAGEMENT_SHEETS)
for i, (mgmt, sn, sheet_num) in enumerate(MANAGEMENT_SHEETS):
    r = 8 + i
    if r > 9:
        copy_from_ref(tmpl0, 9, ws0, r, mc=13)
    ws0.cell(r, 3).value = i + 1 if i == 0 else f"=C{r-1}+1"
    ws0.cell(r, 3).font  = DATA_FONT
    ws0.cell(r, 5).value = "=B5" if i == 0 else None
    ws0.cell(r, 6).value = "-"
    ws0.cell(r, 7).value = mgmt;       ws0.cell(r, 7).font = DATA_FONT
    ws0.cell(r, 8).value = sheet_num;  ws0.cell(r, 8).font = DATA_FONT
    set_h(ws0, r, row_h(mgmt))

# Footer sheet 0
all_r = 8 + n
copy_from_ref(tmpl0, 11, ws0, all_r, mc=13)
ws0.cell(all_r, 3).value = "All"; ws0.cell(all_r, 3).font = DATA_FONT
set_h(ws0, all_r, 18)

copy_from_ref(tmpl0, 12, ws0, all_r + 1, mc=13)

com_r = all_r + 2
copy_from_ref(tmpl0, 13, ws0, com_r, mc=13)
ws0.cell(com_r, 2).value = "Common"; ws0.cell(com_r, 2).font = DATA_FONT
ws0.cell(com_r, 3).value = "-";      ws0.cell(com_r, 3).font = DATA_FONT
set_h(ws0, com_r, 18)

copy_from_ref(tmpl0, 14, ws0, com_r + 1, mc=13)
copy_from_ref(tmpl0, 15, ws0, com_r + 2, mc=13)

for r in range(com_r + 3, 35):
    wipe_row(ws0, r, mc=13)

# ── Detail sheets ──────────────────────────────────────────────────────────────
for mgmt, sn, _sheet_num in MANAGEMENT_SHEETS:
    ws = wb[sn] if sn in wb.sheetnames else wb.copy_worksheet(wb["1"])
    ws.title = sn
    ws.sheet_view.showGridLines = False

    ws["B2"].value     = mgmt
    ws["B2"].font      = TITLE_FONT
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=0)

    # Header integers (row 4)
    for col, val in {2: 1, 3: 2, 5: 1, 6: 2, 7: 3, 8: 4, 9: 5, 11: 1, 12: 2}.items():
        ws.cell(4, col).value = val

    subs   = SUB_ITEMS.get(mgmt, [])
    n_subs = len(subs)

    for i, (name, detail) in enumerate(subs):
        r = 8 + i
        if r > 9:
            copy_from_ref(tmpl1, 9, ws, r)
        ws.cell(r, 3).value = i + 1 if i == 0 else f"=C{r-1}+1"
        ws.cell(r, 3).font  = DATA_FONT
        ws.cell(r, 5).value = "=E5" if i == 0 else None
        ws.cell(r, 7).value = name;   ws.cell(r, 7).font = DATA_FONT
        ws.cell(r, 8).value = detail; ws.cell(r, 8).font = DATA_FONT
        ws.cell(r, 8).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True, indent=1)
        set_h(ws, r, row_h(detail))

    # Footer
    all_r2  = 8 + n_subs
    copy_from_ref(tmpl1, 28, ws, all_r2)
    ws.cell(all_r2, 3).value = "All"; ws.cell(all_r2, 3).font = DATA_FONT
    set_h(ws, all_r2, 18)

    copy_from_ref(tmpl1, 29, ws, all_r2 + 1)

    com_r2 = all_r2 + 2
    copy_from_ref(tmpl1, 30, ws, com_r2)
    ws.cell(com_r2, 2).value = "Common"; ws.cell(com_r2, 2).font = DATA_FONT
    ws.cell(com_r2, 3).value = "-";      ws.cell(com_r2, 3).font = DATA_FONT
    set_h(ws, com_r2, 18)

    close_r = com_r2 + 1
    copy_from_ref(tmpl1, 31, ws, close_r)
    set_h(ws, close_r, 18)

    for r2 in range(close_r + 1, 45):
        wipe_row(ws, r2)

# ── Sheet order: 0 → 1 → 1D → 2 → 2D → 3 → 3D → 4 → 5 → 6 ─────────────────
SHEET_ORDER = ["0", "1", "1D", "2", "2D", "3", "3D", "4", "5", "6"]
for target_i, nm in enumerate(SHEET_ORDER):
    if nm in wb.sheetnames:
        current_i = wb.sheetnames.index(nm)
        offset    = target_i - current_i
        if offset != 0:
            wb.move_sheet(nm, offset=offset)

wb.save(OUTPUT)
print(f"Done: {OUTPUT}")
