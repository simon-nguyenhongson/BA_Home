import openpyxl, math, shutil
from openpyxl.styles import Font, Alignment, Border
from copy import copy
from datetime import date
import os

ROOT_TITLE = "BA_Home — Nền Tảng Quản Trị Dự Án IT"

MANAGEMENT = [
    "PPG System — Quản Trị Dự Án",
    "BA Workflow — Quản Lý Tài Liệu",
    "Test Platform — Tự Động Hóa Kiểm Thử",
    "Kiến Trúc & Hạ Tầng",
    "Luồng Nghiệp Vụ Chính",
    "Vận Hành & Triển Khai",
]

SUB_ITEMS = {
    "PPG System — Quản Trị Dự Án": [
        ("Quản lý dự án & trạng thái",
         "Tạo, cập nhật, archive dự án IT với 4 trạng thái active/on_hold/completed/archived; liên kết dự án với Annual Plan; dashboard KPI tổng hợp số tài liệu, test coverage, pass/fail rate."),
        ("Auto-generate 9 Milestones chuẩn",
         "Khi tạo dự án, hệ thống tự động sinh 9 milestone từ Kickoff đến Closure, phân bổ tỷ lệ theo timeline; mỗi milestone tự động gắn BA tasks và Test tasks tương ứng."),
        ("Quản lý thành viên & hệ thống Alias",
         "Quản lý vai trò PM/BA/Dev/QA/PO/Stakeholder trong dự án; hệ thống @alias để tự động map tên đầy đủ thành viên khi parse meeting notes từ ghi chú thô."),
        ("Quản lý file & File Versioning",
         "Hỗ trợ 3 loại file: template, uploaded, external_url; lưu lịch sử phiên bản đầy đủ mỗi lần upload mới; tải xuống theo version cụ thể; copy file từ URL ngoài vào storage nội bộ."),
        ("Meeting Minutes Parser",
         "Input ghi chú thô với @alias tokens, output biên bản họp có cấu trúc JSON gồm attendees, decisions, action items có assignee và due date, risks, suggestions."),
    ],
    "BA Workflow — Quản Lý Tài Liệu": [
        ("Document State Machine 4 trạng thái",
         "Luồng phê duyệt chuẩn: draft → review → approved → archived; reject quay về draft; lưu snapshot lịch sử mỗi lần chỉnh sửa; đảm bảo traceability toàn bộ vòng đời tài liệu."),
        ("Hỗ trợ 4 loại tài liệu nghiệp vụ",
         "Quản lý BRD (Business Requirements Document), BRS (Business Requirements Specification), ERD (Entity Relationship Diagram), API Specification với versioning đầy đủ cho từng loại."),
        ("Auto-push khi Approve tài liệu",
         "Khi tài liệu được approve, tất cả doc types tự động đẩy sang PPG qua POST /sync-doc; riêng BRS đẩy sang Test Platform qua POST /brs để trigger auto-generate test cases nền."),
        ("Stakeholder Discussions",
         "Tạo thảo luận gắn theo tài liệu cụ thể với 3 trạng thái open/resolved/deferred; ghi nhận resolution notes; làm rõ yêu cầu nghiệp vụ trước khi chuyển sang phê duyệt."),
        ("BA Timeline — Gantt View",
         "Xem tiến độ BA tasks theo milestone dạng Gantt; tracking trạng thái từng tài liệu theo giai đoạn; hỗ trợ BA nắm bắt toàn bộ tiến độ documentation trong dự án."),
    ],
    "Test Platform — Tự Động Hóa Kiểm Thử": [
        ("Auto Test Case Generation từ BRS",
         "Mỗi business rule trong BRS tự động sinh 1 test case + 1 Playwright script E2E; xử lý nền (background task) không chặn workflow; kết quả sẵn sàng cho Tester review ngay."),
        ("Playwright Script Auto-generation",
         "Tạo Playwright E2E script tự động cho từng business rule với template chuẩn; Playwright Script Viewer tích hợp trong UI để xem và copy script; sẵn sàng chạy sau khi bổ sung assertion."),
        ("Test Case State Machine",
         "Luồng kiểm thử 4 bước: generated → reviewed → approved → executed; Tester review và approve từng test case trước khi execute; đảm bảo chất lượng kiểm soát theo quy trình chuẩn."),
        ("Test Report Approval & Sync",
         "Tạo báo cáo kiểm thử thủ công với total/passed/failed; hệ thống tự tính coverage = passed/total×100; approve report → tự động push metrics về PPG cập nhật KPI dashboard dự án."),
        ("Rediff — Tái sinh Test Cases",
         "Khi BRS thay đổi, trigger POST /brs/{id}/rediff để tái sinh test cases phù hợp với yêu cầu mới; đảm bảo test coverage luôn được đồng bộ với phiên bản BRS mới nhất."),
    ],
    "Kiến Trúc & Hạ Tầng": [
        ("Kiến trúc Microservices 3 Services",
         "PPG (:8001), BA Workflow (:8002), Test Platform (:8003) độc lập; giao tiếp qua HTTP REST và sync endpoints; chia sẻ PostgreSQL devops_hub DB; API Proxy Vite tự động route theo prefix."),
        ("Tech Stack Backend",
         "FastAPI 0.115.5 + asyncpg 0.30.0 + Pydantic 2.10.3 làm nền tảng; Python 3.12+; PostgreSQL 15+; aiokafka 0.13.0 cho event streaming; httpx 0.28.1 cho inter-service HTTP calls."),
        ("Frontend React SPA",
         "React 18.3.1 + TypeScript 5.4.5 + Vite 5.2.12 + Zustand 4.5.2;  Design System với StatusBadge, Modal, KpiCard, ProgressBar, Toast; Vite proxy route /api/ppg, /api/ba, /api/test."),
        ("Database & File Storage",
         "PostgreSQL 15 devops_hub dùng chung cho cả 3 services; asyncpg connection pool; file storage local filesystem tại uploads/{project_id}/; upgrade path rõ ràng sang MinIO/S3 khi cần scale."),
        ("Event Streaming & Caching (Optional)",
         "Apache Kafka 3+ cho event streaming giữa các services qua aiokafka consumer; Redis 7+ cho caching; MongoDB 6+ cho document store mở rộng; graceful fallback khi Kafka không khả dụng."),
    ],
    "Luồng Nghiệp Vụ Chính": [
        ("Flow 1: Tạo Dự Án Mới",
         "POST /projects tự động sinh 9 milestones phân bổ theo timeline, tạo thư mục uploads/{id}/, auto-gen BA tasks và Test tasks cho từng milestone, sinh file template markdown Charter và BRD."),
        ("Flow 2: BA Phê Duyệt Tài Liệu",
         "Tạo draft → chỉnh sửa lưu history → submit_review → approve; khi approve: tự động sync tất cả doc sang PPG qua /sync-doc và đẩy BRS sang Test Platform để trigger auto-gen test cases."),
        ("Flow 3: Auto-generate Test Cases",
         "BRS được approve tại BA → POST /brs tới Test Platform → background generate_test_cases_from_brs() parse modules/rules → mỗi rule sinh 1 TestCase + 1 Playwright script sẵn sàng review."),
        ("Flow 4: Báo Cáo Kiểm Thử",
         "Tester tạo report với total/passed/failed; hệ thống tính coverage tự động; approve report → auto push metrics về PPG /sync-test; dashboard dự án cập nhật KPI test coverage realtime."),
        ("Flow 5: Meeting Minutes Parser",
         "POST /meetings/generate với raw_notes chứa @alias tokens; hệ thống map alias thành viên, trích xuất discussion items, decisions, action items (assignee + due date), risks, suggestions dạng JSON."),
    ],
    "Vận Hành & Triển Khai": [
        ("Khởi động & Cài đặt",
         "start.bat mở 4 cửa sổ CMD riêng cho PPG/BA/Test/Frontend; install.bat cài dependencies Python và npm; psql -f infra/init.sql khởi tạo schema; health check GET /health per service."),
        ("Cấu hình môi trường (.env)",
         "DATABASE_URL, KAFKA_BOOTSTRAP, REDIS_URL cấu hình per service; BA service cần PPG_SERVICE_URL=:8001 và TEST_SERVICE_URL=:8003 để auto-push; SECRET_KEY riêng mỗi service."),
        ("Known Issues & Giới hạn (6 issues)",
         "Kafka không kết nối local (Low); Playwright template cần implement assertion (Medium); Vite CJS warning (Low); copy-from-URL thiếu timeout (Low); alias parser phụ thuộc member; thiếu authentication (High)."),
        ("Lịch sử phiên bản — 3 Releases",
         "v1.0.0 MVP monolith PostgreSQL; v2.0.0 microservices 3 services + auto-gen milestones + BRS sync pipeline; v3.0.0 full features: versioning, state machine, Playwright auto-gen,  Design System."),
        ("Yêu cầu hệ thống & môi trường",
         "Bắt buộc: Python 3.12+, Node.js 20+ / npm 10+, PostgreSQL 15+; tùy chọn: Kafka 3+, Redis 7+, MongoDB 6+; hỗ trợ Windows start.bat và Linux/Mac manual run với uvicorn --reload."),
    ],
}

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "GNM.xlsx")
OUTPUT = os.path.normpath(os.path.join(
    os.path.dirname(__file__), "..",
    f"gnm_ba_home_{date.today().strftime('%Y%m%d')}.xlsx"
))

FONT_NAME  = "Myriad Pro"
DATA_FONT  = Font(name=FONT_NAME, size=11)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True)
EMPTY_BDR  = Border()


def copy_from_ref(src, sr, dst, dr, mc=14):
    for col in range(1, mc + 1):
        s = src.cell(sr, col); d = dst.cell(dr, col)
        d.value = s.value
        if s.has_style:
            d.font       = copy(s.font)
            d.fill       = copy(s.fill)
            d.border     = copy(s.border)
            d.alignment  = copy(s.alignment)
            d.number_format = s.number_format


def wipe_row(ws, r, mc=14):
    for col in range(1, mc + 1):
        c = ws.cell(r, col); c.value = None; c.border = EMPTY_BDR


def set_h(ws, r, h):
    ws.row_dimensions[r].height = h


def row_h(txt, base=18):
    return max(base, math.ceil(len(str(txt or "")) / 100) * base + 4)


# Load template as read-only reference
tmpl_ref = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
tmpl0    = tmpl_ref["0"]
tmpl1    = tmpl_ref["1"]

shutil.copy(TEMPLATE_PATH, OUTPUT)
wb = openpyxl.load_workbook(OUTPUT)

# ── Sheet 0 ───────────────────────────────────────────────────────────────────
ws0 = wb["0"]
ws0.sheet_view.showGridLines = False

ws0["B2"].value     = ROOT_TITLE
ws0["B2"].font      = TITLE_FONT
ws0["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=0)

for col, val in {2: 1, 3: 2, 5: 1, 6: 2, 7: 3, 8: 4, 10: 1, 11: 2}.items():
    ws0.cell(4, col).value = val

n = len(MANAGEMENT)
for i, mgmt in enumerate(MANAGEMENT):
    r = 8 + i
    if r > 9:
        copy_from_ref(tmpl0, 9, ws0, r, mc=13)
    ws0.cell(r, 3).value = i + 1 if i == 0 else f"=C{r-1}+1"
    ws0.cell(r, 3).font  = DATA_FONT
    ws0.cell(r, 5).value = "=B5" if i == 0 else None
    ws0.cell(r, 6).value = "-"
    ws0.cell(r, 7).value = mgmt;  ws0.cell(r, 7).font = DATA_FONT
    ws0.cell(r, 8).value = i + 1; ws0.cell(r, 8).font = DATA_FONT
    set_h(ws0, r, row_h(mgmt))

all_r = 8 + n
copy_from_ref(tmpl0, 11, ws0, all_r, mc=13)
ws0.cell(all_r, 3).value = "All"; ws0.cell(all_r, 3).font = DATA_FONT
set_h(ws0, all_r, 18)

copy_from_ref(tmpl0, 12, ws0, all_r + 1, mc=13)

com_r = all_r + 2
copy_from_ref(tmpl0, 13, ws0, com_r, mc=13)
ws0.cell(com_r, 2).value = "Common"; ws0.cell(com_r, 2).font = DATA_FONT
ws0.cell(com_r, 3).value = "-";      ws0.cell(com_r, 3).font = DATA_FONT
set_h(ws0, com_r, 18)

copy_from_ref(tmpl0, 14, ws0, com_r + 1, mc=13)
copy_from_ref(tmpl0, 15, ws0, com_r + 2, mc=13)

for r in range(com_r + 3, 30):
    wipe_row(ws0, r, mc=13)

# ── Detail sheets 1..N ───────────────────────────────────────────────────────
for idx, mgmt in enumerate(MANAGEMENT):
    sn = str(idx + 1)
    ws = wb[sn] if sn in wb.sheetnames else wb.copy_worksheet(wb["1"])
    ws.title = sn
    ws.sheet_view.showGridLines = False

    ws["B2"].value     = mgmt
    ws["B2"].font      = TITLE_FONT
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=0)

    for col, val in {2: 1, 3: 2, 5: 1, 6: 2, 7: 3, 8: 4, 9: 5, 11: 1, 12: 2}.items():
        ws.cell(4, col).value = val

    subs   = SUB_ITEMS.get(mgmt, [])
    n_subs = len(subs)

    for i, (name, detail) in enumerate(subs):
        r = 8 + i
        if r > 9:
            copy_from_ref(tmpl1, 9, ws, r)
        ws.cell(r, 3).value = i + 1 if i == 0 else f"=C{r-1}+1"
        ws.cell(r, 3).font  = DATA_FONT
        ws.cell(r, 5).value = "=E5" if i == 0 else None
        ws.cell(r, 7).value = name;   ws.cell(r, 7).font = DATA_FONT
        ws.cell(r, 8).value = detail; ws.cell(r, 8).font = DATA_FONT
        ws.cell(r, 8).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True, indent=1)
        set_h(ws, r, row_h(detail))

    all_r2 = 8 + n_subs
    copy_from_ref(tmpl1, 28, ws, all_r2)
    ws.cell(all_r2, 3).value = "All"; ws.cell(all_r2, 3).font = DATA_FONT
    set_h(ws, all_r2, 18)

    copy_from_ref(tmpl1, 29, ws, all_r2 + 1)

    com_r2 = all_r2 + 2
    copy_from_ref(tmpl1, 30, ws, com_r2)
    ws.cell(com_r2, 2).value = "Common"; ws.cell(com_r2, 2).font = DATA_FONT
    ws.cell(com_r2, 3).value = "-";      ws.cell(com_r2, 3).font = DATA_FONT
    set_h(ws, com_r2, 18)

    close_r = com_r2 + 1
    copy_from_ref(tmpl1, 31, ws, close_r)
    set_h(ws, close_r, 18)

    for r2 in range(close_r + 1, 40):
        wipe_row(ws, r2)

# Sắp xếp sheets đúng thứ tự
for i, nm in enumerate(["0", "1", "2", "3", "4", "5", "6", "7", "8"]):
    if nm in wb.sheetnames:
        wb.move_sheet(nm, offset=i - wb.sheetnames.index(nm))

wb.save(OUTPUT)
print(f"Done: {OUTPUT}")
